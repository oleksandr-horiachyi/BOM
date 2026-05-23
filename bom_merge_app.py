"""
BOM Merge Tool — Streamlit Web Version
FabTools | Tool 2
Merges multiple BOM Excel files into one, tracking changes with color highlights.
"""

import io
import zipfile
from copy import copy
from typing import Any, Dict, List, Optional, Tuple

import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="BOM Merge Tool — FabTools",
    page_icon="🔀",
    layout="centered",
)

# ─────────────────────────────────────────────────────────────
# CUSTOM CSS
# ─────────────────────────────────────────────────────────────

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Source+Sans+3:wght@400;500;600;700&display=swap');
html, body, [class*="css"] { font-family: 'Source Sans 3', sans-serif; }

.hero-banner {
    background: linear-gradient(135deg, #0C447C 0%, #185FA5 55%, #378ADD 100%);
    border-radius: 12px; padding: 36px 36px 28px; color: white; margin-bottom: 32px;
}
.hero-banner h1 { font-size: 28px; font-weight: 700; margin: 0 0 8px 0; line-height: 1.2; }
.hero-banner p { font-size: 15px; color: rgba(255,255,255,0.82); margin: 0 0 20px 0; line-height: 1.6; }
.hero-steps { display: flex; align-items: center; gap: 10px; flex-wrap: wrap; }
.hero-step { display: flex; align-items: center; gap: 8px; font-size: 14px; font-weight: 600; color: rgba(255,255,255,0.95); }
.step-num {
    width: 24px; height: 24px; border-radius: 50%;
    background: rgba(255,255,255,0.2); border: 1.5px solid rgba(255,255,255,0.4);
    display: flex; align-items: center; justify-content: center;
    font-size: 12px; font-weight: 700; text-align: center; line-height: 24px;
}
.hero-arrow { color: rgba(255,255,255,0.4); font-size: 16px; }
.trust-badge {
    display: inline-flex; align-items: center; gap: 8px;
    background: rgba(255,255,255,0.1); border: 1px solid rgba(255,255,255,0.2);
    border-radius: 8px; padding: 7px 14px; font-size: 13px;
    color: rgba(255,255,255,0.85); margin-top: 16px;
}
.section-header {
    font-size: 13px; font-weight: 600; letter-spacing: 1.2px;
    text-transform: uppercase; color: #185FA5; margin: 0 0 12px 0;
    padding-top: 8px; border-top: 2px solid #E6F1FB;
}
.info-box {
    background: #E6F1FB; border-left: 3px solid #185FA5;
    border-radius: 0 8px 8px 0; padding: 10px 14px;
    font-size: 13px; color: #0C447C; line-height: 1.55; margin-bottom: 12px;
}
.file-order-card {
    background: #F8F9FC; border: 1px solid #DDE1EA;
    border-radius: 8px; padding: 10px 14px; margin-bottom: 8px;
    display: flex; align-items: center; gap: 10px;
}
.file-badge {
    display: inline-flex; align-items: center; gap: 6px;
    padding: 3px 10px; border-radius: 4px; font-size: 12px; font-weight: 600;
}
.badge-base { background: #185FA5; color: white; }
.badge-merge { background: #E6F1FB; color: #185FA5; border: 1px solid #185FA5; }
.color-swatch {
    display: inline-block; width: 14px; height: 14px;
    border-radius: 3px; border: 1px solid rgba(0,0,0,0.15);
    vertical-align: middle; margin-right: 4px;
}
.stats-row {
    display: grid; grid-template-columns: repeat(3, 1fr); gap: 12px; margin-bottom: 16px;
}
.stat-card {
    background: #F8F9FC; border: 1px solid #DDE1EA; border-radius: 8px;
    padding: 14px; text-align: center;
}
.stat-num { font-size: 28px; font-weight: 700; color: #185FA5; }
.stat-label { font-size: 12px; color: #8E96A8; margin-top: 2px; }
.help-btn {
    position: fixed; bottom: 24px; left: 24px; z-index: 999;
    background: #185FA5; color: white !important;
    width: 46px; height: 46px; border-radius: 50%;
    display: flex; align-items: center; justify-content: center;
    font-size: 20px; text-decoration: none !important;
    box-shadow: 0 4px 16px rgba(24,95,165,0.35);
    transition: transform 0.15s, box-shadow 0.15s;
}
.help-btn:hover { transform: translateY(-2px); box-shadow: 0 6px 22px rgba(24,95,165,0.5); }

.footer-bar {
    text-align: center; font-size: 12px; color: #888;
    margin-top: 40px; padding-top: 20px; border-top: 1px solid #EFF1F5;
}
.change-added { color: #1D9E75; font-weight: 600; }
.change-summed { color: #185FA5; font-weight: 600; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────
# MERGE COLORS — one per file merged (skip file 1 = base)
# ─────────────────────────────────────────────────────────────

MERGE_COLORS = [
    ("FFF2CC", "🟡 Yellow"),
    ("CCE5FF", "🔵 Blue"),
    ("D5E8D4", "🟢 Green"),
    ("FAD7E6", "🩷 Pink"),
    ("E8E7F6", "🟣 Purple"),
    ("FFD8B1", "🟠 Orange"),
    ("D6EAF8", "🩵 Cyan"),
    ("FDEBD0", "🍑 Peach"),
]

# ─────────────────────────────────────────────────────────────
# CORE MERGE FUNCTIONS
# ─────────────────────────────────────────────────────────────

def last_used_row(ws) -> int:
    """Find last row with actual data (columns 1-5)."""
    for r in range(ws.max_row, 2, -1):
        if any(ws.cell(r, c).value not in (None, "") for c in range(1, 6)):
            return r
    return 2


def copy_row_style(ws, src_row: int, dst_row: int) -> None:
    """Copy cell styles from one row to another within same sheet."""
    for c in range(1, 6):
        s = ws.cell(src_row, c)
        d = ws.cell(dst_row, c)
        if s.has_style:
            d.font = copy(s.font)
            d.border = copy(s.border)
            d.fill = copy(s.fill)
            d.number_format = s.number_format
            d.protection = copy(s.protection)
            d.alignment = copy(s.alignment)


def copy_sheet_exact(src, dst) -> None:
    """Copy full sheet content and formatting to a new sheet."""
    for col, dim in src.column_dimensions.items():
        dst.column_dimensions[col].width = dim.width
    for row, dim in src.row_dimensions.items():
        dst.row_dimensions[row].height = dim.height
    for row in src.iter_rows():
        for cell in row:
            new = dst.cell(cell.row, cell.column, cell.value)
            if cell.has_style:
                new.font = copy(cell.font)
                new.border = copy(cell.border)
                new.fill = copy(cell.fill)
                new.number_format = cell.number_format
                new.protection = copy(cell.protection)
                new.alignment = copy(cell.alignment)
    for merged in src.merged_cells.ranges:
        dst.merge_cells(str(merged))


def clean_empty(ws) -> int:
    """Remove rows where REF is empty or QTY is 0. Returns count removed."""
    removed = 0
    r = 3
    while r <= ws.max_row:
        ref = ws.cell(r, 1).value
        qty = ws.cell(r, 3).value
        is_empty = ref is None or str(ref).strip() == ""
        is_zero = qty == 0
        if is_empty or is_zero:
            ws.delete_rows(r, 1)
            removed += 1
        else:
            r += 1
    return removed


def rewrite_formulas(ws) -> None:
    """Recalculate column E formula for all data rows."""
    for r in range(3, ws.max_row + 1):
        if ws.cell(r, 1).value not in (None, ""):
            ws.cell(r, 5).value = f"=C{r}*$E$2"


def merge_one_file(
    base_wb,
    merge_wb,
    color_hex: str,
    file_label: str,
) -> List[Dict]:
    """
    Merge merge_wb into base_wb in place.
    Returns list of change records.
    """
    changes = []
    fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

    for sheet_name in merge_wb.sheetnames:
        ws2 = merge_wb[sheet_name]

        # Sheet doesn't exist in base → copy entirely
        if sheet_name not in base_wb.sheetnames:
            ws_new = base_wb.create_sheet(sheet_name)
            copy_sheet_exact(ws2, ws_new)
            # Highlight all data rows as new (added from this file)
            for r in range(3, ws_new.max_row + 1):
                if ws_new.cell(r, 1).value not in (None, ""):
                    for c in range(1, 6):
                        ws_new.cell(r, c).fill = fill
            changes.append({
                "file": file_label,
                "sheet": sheet_name,
                "ref": "— (entire sheet)",
                "action": "SHEET ADDED",
                "old_qty": None,
                "new_qty": None,
            })
            continue

        # Sheet exists → merge row by row
        ws1 = base_wb[sheet_name]
        end = last_used_row(ws1)
        template_row = max(end, 3)

        # Build index: ref_code → row number in base
        ref_index: Dict[str, int] = {}
        for r in range(3, end + 1):
            ref_val = ws1.cell(r, 1).value
            if ref_val not in (None, ""):
                ref_index[str(ref_val).strip()] = r

        for r2 in range(3, ws2.max_row + 1):
            ref_val = ws2.cell(r2, 1).value
            if ref_val is None or str(ref_val).strip() == "":
                continue
            ref = str(ref_val).strip()
            qty2 = ws2.cell(r2, 3).value or 0
            if qty2 == 0:
                continue

            if ref in ref_index:
                # REF exists — sum QTY
                r1 = ref_index[ref]
                old_qty = ws1.cell(r1, 3).value or 0
                new_qty = old_qty + qty2
                ws1.cell(r1, 3).value = new_qty
                ws1.cell(r1, 3).fill = fill  # highlight only QTY cell
                changes.append({
                    "file": file_label,
                    "sheet": sheet_name,
                    "ref": ref,
                    "action": "SUMMED",
                    "old_qty": old_qty,
                    "new_qty": new_qty,
                })
            else:
                # REF is new — add row
                end += 1
                ws1.insert_rows(end)
                copy_row_style(ws1, template_row, end)
                ws1.cell(end, 1).value = ref
                ws1.cell(end, 2).value = ws2.cell(r2, 2).value
                ws1.cell(end, 3).value = qty2
                ws1.cell(end, 4).value = ws2.cell(r2, 4).value
                ws1.cell(end, 5).value = f"=C{end}*$E$2"
                for c in range(1, 6):
                    ws1.cell(end, c).fill = fill
                ref_index[ref] = end
                changes.append({
                    "file": file_label,
                    "sheet": sheet_name,
                    "ref": ref,
                    "action": "ADDED",
                    "old_qty": None,
                    "new_qty": qty2,
                })

    return changes


def merge_all_files(file_bytes_list: List[Tuple[str, bytes]]) -> Tuple[bytes, bytes, List[Dict]]:
    """
    Merge all BOM files sequentially.
    First file = base. Each subsequent file is merged into the running result.
    Returns: (merged_xlsx_bytes, report_xlsx_bytes, all_changes)
    """
    # Load base workbook
    base_name, base_bytes = file_bytes_list[0]
    base_wb = load_workbook(io.BytesIO(base_bytes))

    all_changes: List[Dict] = []

    # Merge each subsequent file
    for idx, (fname, fbytes) in enumerate(file_bytes_list[1:], start=1):
        color_hex = MERGE_COLORS[(idx - 1) % len(MERGE_COLORS)][0]
        merge_wb = load_workbook(io.BytesIO(fbytes))
        changes = merge_one_file(base_wb, merge_wb, color_hex, fname)
        all_changes.extend(changes)

    # Clean and fix formulas in all sheets
    for ws in base_wb.worksheets:
        clean_empty(ws)
        rewrite_formulas(ws)

    # Save merged workbook to bytes
    merged_buf = io.BytesIO()
    base_wb.save(merged_buf)
    merged_bytes = merged_buf.getvalue()

    # Build report workbook
    report_wb = Workbook()
    ws_rep = report_wb.active
    ws_rep.title = "Merge Report"

    # Header
    headers = ["#", "Source File", "Sheet", "REF Code", "Action", "Old QTY", "New QTY", "Change"]
    for col, h in enumerate(headers, 1):
        c = ws_rep.cell(1, col, h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9EAF7")
        c.alignment = Alignment(horizontal="center")

    # Data rows
    for i, ch in enumerate(all_changes, start=2):
        action = ch["action"]
        old = ch["old_qty"]
        new = ch["new_qty"]
        change = ""
        if action == "SUMMED" and old is not None and new is not None:
            change = f"+{new - old}"

        row_data = [i - 1, ch["file"], ch["sheet"], ch["ref"], action, old, new, change]
        for col, val in enumerate(row_data, 1):
            c = ws_rep.cell(i, col, val)
            c.alignment = Alignment(horizontal="center" if col != 4 else "left")

        # Color rows by action
        row_fill = None
        if action == "ADDED":
            row_fill = PatternFill("solid", fgColor="D5E8D4")
        elif action == "SUMMED":
            row_fill = PatternFill("solid", fgColor="D9EAF7")
        elif action == "SHEET ADDED":
            row_fill = PatternFill("solid", fgColor="FFF2CC")
        if row_fill:
            for col in range(1, 9):
                ws_rep.cell(i, col).fill = row_fill

    # Auto width
    for col in ws_rep.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws_rep.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)

    ws_rep.auto_filter.ref = f"A1:H{ws_rep.max_row}"
    ws_rep.freeze_panes = "A2"

    # Summary sheet
    ws_sum = report_wb.create_sheet("Summary")
    total_added = sum(1 for ch in all_changes if ch["action"] == "ADDED")
    total_summed = sum(1 for ch in all_changes if ch["action"] == "SUMMED")
    total_sheets = sum(1 for ch in all_changes if ch["action"] == "SHEET ADDED")
    files_merged = len(file_bytes_list) - 1

    sum_data = [
        ["Merge Summary", ""],
        ["", ""],
        ["Files merged", files_merged],
        ["Rows added (new REF codes)", total_added],
        ["Rows updated (QTY summed)", total_summed],
        ["Sheets added from merge files", total_sheets],
        ["Total changes", len(all_changes)],
    ]
    for row_data in sum_data:
        ws_sum.append(row_data)
    ws_sum["A1"].font = Font(bold=True, size=14)
    ws_sum.column_dimensions["A"].width = 35
    ws_sum.column_dimensions["B"].width = 15

    report_buf = io.BytesIO()
    report_wb.save(report_buf)
    report_bytes = report_buf.getvalue()

    return merged_bytes, report_bytes, all_changes


# ─────────────────────────────────────────────────────────────
# UI
# ─────────────────────────────────────────────────────────────

# Hero
st.markdown("""
<div class="hero-banner">
  <h1>🔀 BOM Merge Tool</h1>
  <p>Upload multiple BOM Excel files and merge them into one.
     QTY values are summed for matching part codes, new parts are added automatically.
     Every change is highlighted by color and logged in a report.</p>
  <div class="hero-steps">
    <div class="hero-step"><span class="step-num">1</span> Upload files</div>
    <span class="hero-arrow">→</span>
    <div class="hero-step"><span class="step-num">2</span> Set merge order</div>
    <span class="hero-arrow">→</span>
    <div class="hero-step"><span class="step-num">3</span> Run</div>
    <span class="hero-arrow">→</span>
    <div class="hero-step"><span class="step-num">4</span> Download result</div>
  </div>
  <div class="trust-badge">🔒 Your files are processed in memory and never stored on any server.</div>
</div>
""", unsafe_allow_html=True)

# ── STEP 1: Upload ───────────────────────────────────────────
st.markdown('<div class="section-header">Step 1 — Upload BOM files</div>', unsafe_allow_html=True)

st.markdown("""
<div class="info-box">
  Upload <strong>2 or more</strong> BOM Excel files (.xlsx). 
  The <strong>first file</strong> you uploaded becomes the <strong>base</strong> — 
  all other files are merged into it, in order.
  <br><br>
  <strong>Merge rules:</strong>
  matching REF code → QTY is summed &nbsp;|&nbsp;
  new REF code → row is added &nbsp;|&nbsp;
  new sheet → entire sheet is copied
</div>
""", unsafe_allow_html=True)

uploaded_files = st.file_uploader(
    "Drop BOM files here or click to browse",
    type=["xlsx"],
    accept_multiple_files=True,
    label_visibility="collapsed",
)

# ── STEP 2: File order ───────────────────────────────────────
if uploaded_files and len(uploaded_files) >= 2:
    st.markdown('<div class="section-header">Step 2 — Merge order</div>', unsafe_allow_html=True)

    st.markdown("""
    <div class="info-box">
      Files are merged in the order shown below.
      Use the number inputs to reorder them if needed.
      <strong>File #1 is always the base.</strong>
    </div>
    """, unsafe_allow_html=True)

    # Allow user to set custom order
    order_inputs = {}
    for i, f in enumerate(uploaded_files):
        col1, col2 = st.columns([1, 6])
        with col1:
            order_inputs[f.name] = st.number_input(
                f"Order for {f.name}", min_value=1, max_value=len(uploaded_files),
                value=i + 1, key=f"order_{f.name}", label_visibility="collapsed"
            )
        with col2:
            badge = "badge-base" if i == 0 else "badge-merge"
            label = "BASE" if i == 0 else f"Merge #{i}"
            color_info = MERGE_COLORS[(i - 1) % len(MERGE_COLORS)][1] if i > 0 else ""
            color_hex = MERGE_COLORS[(i - 1) % len(MERGE_COLORS)][0] if i > 0 else None
            swatch = f'<span class="color-swatch" style="background:#{color_hex}"></span>' if color_hex else ""
            st.markdown(
                f'<div class="file-order-card">'
                f'<span class="file-badge {badge}">{label}</span>'
                f'📄 {f.name} &nbsp; {swatch}{color_info}'
                f'</div>',
                unsafe_allow_html=True
            )

    # Sort files by user-specified order
    sorted_files = sorted(uploaded_files, key=lambda f: order_inputs[f.name])

    st.markdown("**Merge sequence:**")
    seq_parts = []
    for i, f in enumerate(sorted_files):
        if i == 0:
            seq_parts.append(f"**{f.name}** (base)")
        else:
            color = MERGE_COLORS[(i - 1) % len(MERGE_COLORS)][1]
            seq_parts.append(f"{f.name} {color}")
    st.caption(" → ".join(seq_parts))

elif uploaded_files and len(uploaded_files) == 1:
    st.warning("⚠️ Please upload at least 2 files to merge.")
    sorted_files = []
else:
    sorted_files = []

# ── STEP 3: Run ──────────────────────────────────────────────
st.markdown('<div class="section-header">Step 3 — Run</div>', unsafe_allow_html=True)

st.markdown("""
<div style="margin-bottom:16px">
  <div style="background:#F0F9F5;border:1px solid #A8D5BF;border-radius:8px;padding:14px 18px;">
    <div style="font-size:14px;color:#0A5C3A;margin-bottom:8px;font-weight:700;">📦 Output files you will receive:</div>
    <div style="font-size:13px;color:#3D8A6B;">
      📗 <strong>Merged BOM.xlsx</strong> — combined workbook with all parts and color highlights<br>
      📊 <strong>Merge Report.xlsx</strong> — full log of every change (added rows, updated QTY, new sheets)<br>
      📦 <strong>Both files in a .zip</strong>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

run_ready = len(sorted_files) >= 2
run_clicked = st.button("⚡  Run Merge", type="primary", use_container_width=True, disabled=not run_ready)

if not run_ready and not uploaded_files:
    st.caption("Upload at least 2 BOM files to enable Run.")

# ── STEP 4: Results ──────────────────────────────────────────
if run_clicked and run_ready:
    with st.spinner("Merging files..."):
        try:
            file_bytes_list = [(f.name, f.read()) for f in sorted_files]
            merged_bytes, report_bytes, all_changes = merge_all_files(file_bytes_list)

            # Stats
            total_added   = sum(1 for ch in all_changes if ch["action"] == "ADDED")
            total_summed  = sum(1 for ch in all_changes if ch["action"] == "SUMMED")
            total_sheets  = sum(1 for ch in all_changes if ch["action"] == "SHEET ADDED")

            st.success(f"✅ Merge complete! {len(sorted_files)} files merged.")

            st.markdown(f"""
            <div class="stats-row">
              <div class="stat-card">
                <div class="stat-num">{total_added}</div>
                <div class="stat-label">Rows added</div>
              </div>
              <div class="stat-card">
                <div class="stat-num">{total_summed}</div>
                <div class="stat-label">QTY updated</div>
              </div>
              <div class="stat-card">
                <div class="stat-num">{total_sheets}</div>
                <div class="stat-label">Sheets added</div>
              </div>
            </div>
            """, unsafe_allow_html=True)

            st.markdown('<div class="section-header">Step 4 — Download results</div>', unsafe_allow_html=True)

            # ZIP both files
            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                zf.writestr("Merged BOM.xlsx", merged_bytes)
                zf.writestr("Merge Report.xlsx", report_bytes)
            zip_buf.seek(0)

            col1, col2, col3 = st.columns(3)
            with col1:
                st.download_button(
                    label="📗 Merged BOM.xlsx",
                    data=merged_bytes,
                    file_name="Merged BOM.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            with col2:
                st.download_button(
                    label="📊 Merge Report.xlsx",
                    data=report_bytes,
                    file_name="Merge Report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            with col3:
                st.download_button(
                    label="📦 Download both (.zip)",
                    data=zip_buf.getvalue(),
                    file_name="BOM_Merge_Results.zip",
                    mime="application/zip",
                    use_container_width=True,
                )

            # Change log preview
            if all_changes:
                with st.expander(f"📋 Change log ({len(all_changes)} changes)"):
                    col_headers = ["Source File", "Sheet", "REF Code", "Action", "Old QTY", "New QTY"]
                    rows = []
                    for ch in all_changes:
                        rows.append({
                            "Source File": ch["file"],
                            "Sheet": ch["sheet"],
                            "REF Code": ch["ref"],
                            "Action": ch["action"],
                            "Old QTY": ch["old_qty"] if ch["old_qty"] is not None else "—",
                            "New QTY": ch["new_qty"] if ch["new_qty"] is not None else "—",
                        })
                    import pandas as pd
                    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

        except Exception as e:
            st.error(f"❌ Error during merge: {e}")
            st.info("Please check that all uploaded files are valid BOM Excel files with the expected format.")

# Color legend
with st.expander("🎨 Color coding guide"):
    st.markdown("Each merge file gets a unique highlight color in the output:")
    for i, (hex_color, label) in enumerate(MERGE_COLORS):
        st.markdown(
            f'<span class="color-swatch" style="background:#{hex_color}; width:18px; height:18px; border-radius:4px; display:inline-block; border:1px solid #ccc; vertical-align:middle;"></span>'
            f' &nbsp; File #{i+2} — {label}',
            unsafe_allow_html=True
        )

# Footer
st.markdown('''<a href="https://oleksandr-horiachyi.github.io/BOM/help.html" target="_blank" class="help-btn" title="Help &amp; Documentation">❓</a>''', unsafe_allow_html=True)

st.markdown("""
<div class="footer-bar">
  <a href="https://oleksandr-horiachyi.github.io/BOM/" target="_blank">← Back to FabTools</a>
  &nbsp;·&nbsp;
  <a href="https://oleksandr-horiachyi.github.io/BOM/help.html" target="_blank">📖 Help & Docs</a>
  &nbsp;·&nbsp; Built with Python &amp; Streamlit &nbsp;·&nbsp;
  <a href="https://github.com/oleksandr-horiachyi/BOM" target="_blank">GitHub</a>
</div>
""", unsafe_allow_html=True)
