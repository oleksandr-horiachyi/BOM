"""
BOM Combiner — Streamlit Web Version
FabTools | Tool 3
Step 1: Upload → Step 2: Edit names/qty/delete → Step 3: Drag reorder → Step 4: Generate
"""

import io
from copy import copy
from typing import Dict, List

import streamlit as st
from openpyxl import Workbook, load_workbook
from streamlit_sortables import sort_items

st.set_page_config(page_title="BOM Combiner — FabTools", page_icon="📋", layout="centered")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Source+Sans+3:wght@400;500;600;700&display=swap');
html, body, [class*="css"] { font-family: 'Source Sans 3', sans-serif; }
.hero-banner {
    background: linear-gradient(135deg, #0C447C 0%, #185FA5 55%, #378ADD 100%);
    border-radius: 12px; padding: 32px 36px 24px; color: white; margin-bottom: 28px;
}
.hero-banner h1 { font-size: 26px; font-weight: 700; margin: 0 0 8px 0; }
.hero-banner p  { font-size: 14px; color: rgba(255,255,255,0.82); margin: 0 0 16px 0; line-height: 1.6; }
.hero-steps { display: flex; align-items: center; gap: 8px; flex-wrap: wrap; }
.hero-step  { display: flex; align-items: center; gap: 6px; font-size: 13px; font-weight: 600; color: rgba(255,255,255,0.95); }
.step-num {
    width: 22px; height: 22px; border-radius: 50%;
    background: rgba(255,255,255,0.2); border: 1.5px solid rgba(255,255,255,0.4);
    font-size: 11px; font-weight: 700; display: flex; align-items: center; justify-content: center;
}
.hero-arrow { color: rgba(255,255,255,0.4); font-size: 14px; }
.trust-badge {
    display: inline-flex; align-items: center; gap: 8px;
    background: rgba(255,255,255,0.1); border: 1px solid rgba(255,255,255,0.2);
    border-radius: 8px; padding: 6px 12px; font-size: 12px; color: rgba(255,255,255,0.85); margin-top: 12px;
}
.section-header {
    font-size: 12px; font-weight: 600; letter-spacing: 1.2px; text-transform: uppercase;
    color: #185FA5; margin: 0 0 10px 0; padding-top: 8px; border-top: 2px solid #E6F1FB;
}
.info-box {
    background: #E6F1FB; border-left: 3px solid #185FA5; border-radius: 0 8px 8px 0;
    padding: 9px 12px; font-size: 13px; color: #0C447C; line-height: 1.5; margin-bottom: 10px;
}
.conflict-badge {
    display: inline-block; background: #FEE2E2; color: #B91C1C;
    font-size: 10px; font-weight: 700; padding: 1px 6px; border-radius: 4px;
}
.renamed-badge {
    display: inline-block; background: #FEF3C7; color: #92400E;
    font-size: 10px; font-weight: 700; padding: 1px 6px; border-radius: 4px;
}
.drag-hint { font-size: 12px; color: #8E96A8; margin-bottom: 6px; }
.stats-row { display: flex; gap: 16px; margin-bottom: 14px; flex-wrap: wrap; }
.stat-card { background: #F8F9FC; border: 1px solid #DDE1EA; border-radius: 8px; padding: 12px 20px; text-align: center; }
.stat-num  { font-size: 24px; font-weight: 700; color: #185FA5; }
.stat-label{ font-size: 11px; color: #8E96A8; margin-top: 2px; }
.help-btn {
    position: fixed; bottom: 24px; left: 24px; z-index: 999;
    background: #185FA5; color: white !important;
    width: 46px; height: 46px; border-radius: 50%;
    display: flex; align-items: center; justify-content: center;
    font-size: 20px; text-decoration: none !important;
    box-shadow: 0 4px 16px rgba(24,95,165,0.35);
    transition: transform 0.15s, box-shadow 0.15s;
}
.help-btn:hover { transform: translateY(-2px); box-shadow: 0 6px 22px rgba(24,95,165,0.5); }
.footer-bar{ text-align: center; font-size: 12px; color: #888; margin-top: 40px; padding-top: 20px; border-top: 1px solid #EFF1F5; }
</style>
""", unsafe_allow_html=True)

# ── HELPERS ──────────────────────────────────────────────────

def read_bom_sheets(file_bytes: bytes, filename: str) -> List[Dict]:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    result = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        raw = ws["E2"].value
        try:
            qty = int(float(str(raw))) if raw is not None else 1
        except:
            qty = 1
        result.append({"original_name": sheet_name, "source_file": filename,
                        "original_qty": qty, "file_bytes": file_bytes})
    return result


def copy_sheet_to_wb(src_bytes, src_name, dst_wb, dst_name, new_qty):
    src_wb = load_workbook(io.BytesIO(src_bytes), data_only=False)
    src_ws = src_wb[src_name]
    dst_ws = dst_wb.create_sheet(dst_name)
    for col, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col].width = dim.width
    for row, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row].height = dim.height
    for row in src_ws.iter_rows():
        for cell in row:
            nc = dst_ws.cell(cell.row, cell.column, cell.value)
            if cell.has_style:
                nc.font = copy(cell.font); nc.border = copy(cell.border)
                nc.fill = copy(cell.fill); nc.number_format = cell.number_format
                nc.protection = copy(cell.protection); nc.alignment = copy(cell.alignment)
    for merged in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged))
    dst_ws["E2"].value = new_qty
    if src_name != dst_name:
        a1 = dst_ws["A1"]
        if a1.value and src_name in str(a1.value):
            a1.value = str(a1.value).replace(src_name, dst_name)


def build_combined_bom(rows: List[Dict]) -> bytes:
    dst_wb = Workbook()
    dst_wb.remove(dst_wb.active)
    for row in rows:
        copy_sheet_to_wb(row["file_bytes"], row["original_name"],
                         dst_wb, row["final_name"], row["qty"])
    buf = io.BytesIO()
    dst_wb.save(buf)
    return buf.getvalue()


def make_drag_labels(order: list) -> tuple:
    """
    Build clean labels (no hidden characters) and a label→idx mapping dict.
    Names are unique (enforced in Step 2), so label collision is impossible.
    """
    labels = []
    label_to_idx = {}
    for idx in order:
        name  = st.session_state.names[idx]
        src   = st.session_state.sheets[idx]["source_file"]
        qty   = st.session_state.qtys[idx]
        label = f"{name}  ×{qty}  —  {src}"
        labels.append(label)
        label_to_idx[label] = idx
    return labels, label_to_idx


def init_state(sheets: List[Dict]):
    st.session_state.sheets  = sheets
    st.session_state.order   = list(range(len(sheets)))
    st.session_state.names   = {i: s["original_name"] for i, s in enumerate(sheets)}
    st.session_state.qtys    = {i: s["original_qty"]  for i, s in enumerate(sheets)}
    st.session_state.deleted = {i: False for i in range(len(sheets))}
    st.session_state.result  = None


# ── SESSION STATE ─────────────────────────────────────────────

for key, val in [("sheets", []), ("order", []), ("names", {}),
                 ("qtys", {}), ("deleted", {}), ("result", None), ("prev_files", [])]:
    if key not in st.session_state:
        st.session_state[key] = val

# ── HERO ─────────────────────────────────────────────────────

st.markdown("""
<div class="hero-banner">
  <h1>📋 BOM Combiner</h1>
  <p>Upload multiple BOM files — edit names and quantities first,
     then drag to set the final sheet order.</p>
  <div class="hero-steps">
    <div class="hero-step"><span class="step-num">1</span> Upload</div>
    <span class="hero-arrow">→</span>
    <div class="hero-step"><span class="step-num">2</span> Edit names & qty</div>
    <span class="hero-arrow">→</span>
    <div class="hero-step"><span class="step-num">3</span> Drag to reorder</div>
    <span class="hero-arrow">→</span>
    <div class="hero-step"><span class="step-num">4</span> Generate</div>
  </div>
  <div class="trust-badge">🔒 Files processed in memory — never stored on any server.</div>
</div>
""", unsafe_allow_html=True)

# ── STEP 1: UPLOAD ────────────────────────────────────────────

st.markdown('<div class="section-header">Step 1 — Upload BOM files</div>', unsafe_allow_html=True)
st.markdown("""
<div class="info-box">
  Upload one or more BOM Excel files (.xlsx). Each sheet = one assembly.
  Quantity is read from cell <strong>E2</strong> of each sheet.
</div>
""", unsafe_allow_html=True)

uploaded_files = st.file_uploader(
    "Drop BOM files here", type=["xlsx"],
    accept_multiple_files=True, label_visibility="collapsed",
)

if uploaded_files:
    file_names = [f.name for f in uploaded_files]
    if file_names != st.session_state.prev_files:
        sheets = []
        for f in uploaded_files:
            sheets.extend(read_bom_sheets(f.read(), f.name))
        init_state(sheets)
        st.session_state.prev_files = file_names

if not st.session_state.sheets:
    st.caption("Upload at least one BOM file to continue.")
    st.stop()

sheets = st.session_state.sheets
n = len(sheets)

# ── STEP 2: EDIT NAMES, QTY, DELETE ──────────────────────────

st.markdown('<div class="section-header">Step 2 — Edit names, quantities & resolve conflicts</div>', unsafe_allow_html=True)
st.markdown("""
<div class="info-box">
  Review all assemblies. <strong>Rename</strong> or <strong>delete 🗑</strong> duplicates here first.
  Duplicate names are highlighted in red — they must be resolved before generating.
  <br>Once names are final, proceed to Step 3 to set the sheet order.
</div>
""", unsafe_allow_html=True)

# Detect conflicts among active rows
active_indices = [i for i in range(n) if not st.session_state.deleted[i]]
name_count: Dict[str, int] = {}
for i in active_indices:
    nm = st.session_state.names[i]
    name_count[nm] = name_count.get(nm, 0) + 1
conflict_names = {nm for nm, cnt in name_count.items() if cnt > 1}

# Table header
h_num, h_name, h_src, h_qty, h_del = st.columns([0.5, 4, 2.5, 1.2, 0.7])
h_num.markdown('<span style="font-size:11px;color:#8E96A8;font-weight:700">#</span>', unsafe_allow_html=True)
h_name.markdown('<span style="font-size:11px;color:#8E96A8;font-weight:700">FINAL NAME</span>', unsafe_allow_html=True)
h_src.markdown('<span style="font-size:11px;color:#8E96A8;font-weight:700">SOURCE FILE</span>', unsafe_allow_html=True)
h_qty.markdown('<span style="font-size:11px;color:#8E96A8;font-weight:700">QTY</span>', unsafe_allow_html=True)
h_del.markdown('<span style="font-size:11px;color:#8E96A8;font-weight:700">DEL</span>', unsafe_allow_html=True)
st.divider()

active_pos = 1
for i in range(n):
    asm     = sheets[i]
    deleted = st.session_state.deleted[i]
    cur_name= st.session_state.names[i]
    orig    = asm["original_name"]
    renamed = cur_name != orig
    conflict= (cur_name in conflict_names) and not deleted

    c_num, c_name, c_src, c_qty, c_del = st.columns([0.5, 4, 2.5, 1.2, 0.7])

    with c_num:
        if deleted:
            st.markdown('<span style="color:#ccc">—</span>', unsafe_allow_html=True)
        else:
            st.markdown(f"**{active_pos}.**")
            active_pos += 1

    with c_name:
        if deleted:
            st.markdown(f'<span style="color:#bbb;text-decoration:line-through">{cur_name}</span>',
                        unsafe_allow_html=True)
        else:
            new_name = st.text_input(
                label=f"name_{i}", value=cur_name,
                key=f"name_input_{i}", label_visibility="collapsed",
                help=f"Original sheet name: {orig}",
            )
            if new_name.strip() and new_name.strip() != cur_name:
                st.session_state.names[i] = new_name.strip()
                st.rerun()

            if conflict:
                st.markdown('<span class="conflict-badge">⚠️ DUPLICATE — rename or delete</span>',
                            unsafe_allow_html=True)
            elif renamed:
                st.markdown(f'<span class="renamed-badge">✏️ renamed from "{orig}"</span>',
                            unsafe_allow_html=True)

    with c_src:
        color = "#bbb" if deleted else "#8E96A8"
        st.markdown(f'<span style="font-size:12px;color:{color}">{asm["source_file"]}</span>',
                    unsafe_allow_html=True)

    with c_qty:
        if deleted:
            st.markdown('<span style="color:#bbb">—</span>', unsafe_allow_html=True)
        else:
            new_qty = st.number_input(
                label=f"qty_{i}", min_value=0, max_value=9999,
                value=st.session_state.qtys[i], step=1,
                key=f"qty_input_{i}", label_visibility="collapsed",
            )
            st.session_state.qtys[i] = new_qty

    with c_del:
        if not deleted:
            if st.button("🗑", key=f"del_{i}", help="Exclude from output"):
                st.session_state.deleted[i] = True
                st.rerun()
        else:
            if st.button("↩", key=f"restore_{i}", help="Restore"):
                st.session_state.deleted[i] = False
                st.rerun()

st.divider()

deleted_count = sum(1 for i in range(n) if st.session_state.deleted[i])
active_indices = [i for i in range(n) if not st.session_state.deleted[i]]
total_qty = sum(st.session_state.qtys[i] for i in active_indices)

c1, c2, c3 = st.columns([5, 2, 1])
info = f"**{len(active_indices)} active** assemblies from **{len(uploaded_files)} file(s)**"
if deleted_count:
    info += f" &nbsp;·&nbsp; <span style='color:#aaa'>{deleted_count} excluded</span>"
c1.markdown(info, unsafe_allow_html=True)
c2.markdown("**Total units:**")
c3.markdown(f"**{total_qty}**")

# ── STEP 3: DRAG TO REORDER (using FINAL names) ───────────────

st.markdown('<div class="section-header">Step 3 — Drag to set sheet order</div>', unsafe_allow_html=True)

# Check for remaining conflicts
name_count2: Dict[str, int] = {}
for i in active_indices:
    nm = st.session_state.names[i]
    name_count2[nm] = name_count2.get(nm, 0) + 1
conflicts_remain = {nm for nm, cnt in name_count2.items() if cnt > 1}

if conflicts_remain:
    st.error(f"❌ Resolve duplicate names first: **{', '.join(conflicts_remain)}**")
    st.stop()

if not active_indices:
    st.warning("⚠️ No assemblies to reorder.")
    st.stop()

st.markdown("""
<div class="info-box">
  All names are now final. <strong>Drag</strong> the items below to set the order
  of sheets in the output file. The order here = the order of tabs in the Combined BOM.
</div>
""", unsafe_allow_html=True)

# Ensure order only contains active (non-deleted) indices
current_order = [i for i in st.session_state.order if not st.session_state.deleted[i]]
# Add any new active indices not yet in order
for i in active_indices:
    if i not in current_order:
        current_order.append(i)
st.session_state.order = current_order

# Build clean drag labels + mapping dict (no hidden characters)
drag_labels, label_to_idx = make_drag_labels(st.session_state.order)

st.markdown('<p class="drag-hint">☰ &nbsp;Drag items to reorder</p>', unsafe_allow_html=True)

sorted_labels = sort_items(drag_labels, direction="vertical", key="drag_reorder")

# Map sorted labels back to original indices using the dict
new_order = [label_to_idx[lbl] for lbl in sorted_labels if lbl in label_to_idx]
if new_order != st.session_state.order:
    st.session_state.order = new_order
    st.rerun()

# Show final order summary below drag list
st.markdown("**Final sheet order:**")
order_parts = []
for pos, i in enumerate(st.session_state.order, 1):
    order_parts.append(f"`{pos}. {st.session_state.names[i]}`")
st.markdown("  →  ".join(order_parts))

# ── STEP 4: GENERATE ─────────────────────────────────────────

st.markdown('<div class="section-header">Step 4 — Generate Combined BOM</div>', unsafe_allow_html=True)
st.markdown("""
<div class="info-box">
  All assemblies combined into one BOM file in the order above.
  <strong>Ready to use directly with Tool 1 — BOM Production Automation.</strong>
</div>
""", unsafe_allow_html=True)

st.markdown(f"""
<div class="stats-row">
  <div class="stat-card"><div class="stat-num">{len(uploaded_files)}</div><div class="stat-label">Files</div></div>
  <div class="stat-card"><div class="stat-num">{len(st.session_state.order)}</div><div class="stat-label">Assemblies</div></div>
  <div class="stat-card"><div class="stat-num">{deleted_count}</div><div class="stat-label">Excluded</div></div>
  <div class="stat-card"><div class="stat-num">{total_qty}</div><div class="stat-label">Total units</div></div>
</div>
""", unsafe_allow_html=True)

if st.button("⚡  Generate Combined BOM", type="primary", use_container_width=True):
    final_rows = [
        {"original_name": sheets[i]["original_name"],
         "final_name":    st.session_state.names[i],
         "file_bytes":    sheets[i]["file_bytes"],
         "qty":           st.session_state.qtys[i]}
        for i in st.session_state.order
    ]
    with st.spinner("Building Combined BOM..."):
        try:
            st.session_state.result = build_combined_bom(final_rows)
            st.success(f"✅ Done! {len(final_rows)} assemblies combined.")
        except Exception as e:
            st.error(f"❌ Error: {e}")

if st.session_state.result:
    st.download_button(
        label="📥 Download Combined BOM.xlsx",
        data=st.session_state.result,
        file_name="Combined BOM.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    with st.expander("📋 Final sheet list"):
        for pos, i in enumerate(st.session_state.order, 1):
            renamed = st.session_state.names[i] != sheets[i]["original_name"]
            tag = f" ✏️ *(was {sheets[i]['original_name']})*" if renamed else ""
            st.markdown(
                f"**{pos}.** `{st.session_state.names[i]}`{tag}"
                f" — qty **{st.session_state.qtys[i]}**"
                f" — *{sheets[i]['source_file']}*"
            )

st.markdown('''<a href="https://oleksandr-horiachyi.github.io/BOM/help.html#tool3" target="_blank" class="help-btn" title="Help &amp; Documentation">❓</a>''', unsafe_allow_html=True)

st.markdown("""
<div class="footer-bar">
  <a href="https://oleksandr-horiachyi.github.io/BOM/" target="_blank">← Back to FabTools</a>
  &nbsp;·&nbsp;
  <a href="https://oleksandr-horiachyi.github.io/BOM/help.html" target="_blank">📖 Help & Docs</a>
  &nbsp;·&nbsp; Built with Python &amp; Streamlit &nbsp;·&nbsp;
  <a href="https://github.com/oleksandr-horiachyi/BOM" target="_blank">GitHub</a>
</div>
""", unsafe_allow_html=True)
