"""
BOM Production Automation v13

Purpose:
1) Read a multi-sheet BOM workbook where each sheet is a separate unit BOM.
2) Create a Combined BOM workbook.
3) Read a LOG file with multiple sheets, for example U-1000, U-2000, U-4000, etc.
4) Let the user select one or more LOG sheets.
5) Create a separate Production Schedule sheet for each selected LOG sheet.
6) Do not apply any hard-coded U-4 / U-7 filter.
7) Remove empty LOG columns before adding QTY.
8) Create a fixed Production Schedule header in rows 1-6.
9) Generate a text-based company logo in cells A1:F3, without using an image file.
10) Add a Drawing Codes sheet. Drawing code = first 6 characters of the part code.

Install:
    pip install openpyxl pillow
    pip install customtkinter

Run:
    python bom_production_automation_v13.py
"""



import os
import sys
import re
import tkinter as tk
import customtkinter as ctk

from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image
from pathlib import Path
from tkinter import filedialog, messagebox
from typing import Any, Dict, List, Optional, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")



# =========================================================
# SETTINGS FOR YOUR FILE FORMAT
# =========================================================

# BOM sheets
BOM_START_ROW = 3          # с какой строки начинается список материалов в BOM
BOM_CODE_COL = 1           # колонка A = REF / CODE
BOM_DESCRIPTION_COL = 2    # колонка B = DESCRIPTION
BOM_QTY_COL = 5            # колонка E = total QTY для этого BOM-листа
BOM_UNIT_QTY_CELL = "E2"   # количество юнитов на BOM-листе

# LOG sheets
LOG_HEADER_ROW = 3         # строка заголовков в LOG file
LOG_START_ROW = 4          # первая строка данных в LOG file
LOG_CODE_COL = 1           # колонка A = Ref. No.

# Production Schedule layout
SCHEDULE_HEADER_ROW = 7    # table header row in Production Schedule
SCHEDULE_DATA_ROW = 8      # first data row in Production Schedule

# Output
NUMBER_FORMAT = '# ##0;-# ##0;-'


# =========================================================
# COMMON FUNCTIONS
# =========================================================

def clean_code(value: Any) -> Optional[str]:
    """Clean a part code. Empty values are returned as None."""
    if value is None:
        return None

    text = str(value).replace("\u00a0", " ").strip().upper()
    text = re.sub(r"\s+", " ", text)

    if text in ("", "NONE", "NAN"):
        return None

    return text

def to_number(value: Any) -> float:
    """Safely convert a value to a number. Return 0 if conversion is not possible."""
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return value

    text = str(value).replace(",", ".").strip()
    if text in ("", "-"):
        return 0

    try:
        return float(text)
    except ValueError:
        return 0

def get_drawing_code(part_code: Any) -> Optional[str]:
    """Return the drawing code for a part code. Example: U-4154-22 -> U-4154."""
    code = clean_code(part_code)
    if code is None:
        return None
    return code[:6]


def safe_sheet_title(title: str, existing: Optional[set] = None) -> str:
    """Create a safe Excel sheet name, max 31 characters."""
    if existing is None:
        existing = set()

    title = re.sub(r"[\\/*?:\[\]]", "-", str(title)).strip()
    title = title[:31] if title else "Sheet"

    base = title
    counter = 1
    while title in existing:
        suffix = f"_{counter}"
        title = base[:31 - len(suffix)] + suffix
        counter += 1

    existing.add(title)
    return title


def auto_width(ws, min_width: int = 8, max_width: int = 45) -> None:
    """Auto-adjust worksheet column widths."""
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = max(min_width, min(max_width, max_len + 2))


def style_header_row(ws, row: int = 1, fill_color: str = "D9EAF7") -> None:
    fill = PatternFill("solid", fgColor=fill_color)
    thin = Side(style="thin", color="B7B7B7")
    for cell in ws[row]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)

def try_insert_logo(ws):

    try:

        # DETERMINE APPLICATION DIRECTORY
        if getattr(sys, 'frozen', False):
            app_dir = sys._MEIPASS
        else:
            app_dir = os.path.dirname(os.path.abspath(__file__))

        logo_path = os.path.join(app_dir, "Logo.png")

        # CHECK FILE EXISTS
        if not os.path.exists(logo_path):
            return

        # READ ORIGINAL IMAGE SIZE
        pil_img = Image.open(logo_path)

        original_width, original_height = pil_img.size

        # TARGET HEIGHT
        # Excel row height uses points, image uses pixels
        # 1 point ≈ 1.333 pixels

        row_height_points = 20
        rows_used = 3

        excel_points_height = row_height_points * rows_used

        # Convert Excel points -> pixels
        target_height = int(excel_points_height * 1.333)

        # KEEP PROPORTIONS
        scale = target_height / original_height

        target_width = int(original_width * scale)

        # INSERT INTO EXCEL
        logo = ExcelImage(logo_path)

        logo.width = target_width
        logo.height = target_height

        # INSERT IMAGE
        ws.add_image(logo, "A1")

    except Exception as e:

        print(f"Logo insert skipped: {e}")

def apply_schedule_row_heights(ws) -> None:
    """Set row heights for a Production Schedule sheet: rows 1-6 = 20, all other used rows = 15."""
    for row_idx in range(1, 7):
        ws.row_dimensions[row_idx].height = 20

    last_row = max(ws.max_row, SCHEDULE_DATA_ROW)
    for row_idx in range(7, last_row + 1):
        ws.row_dimensions[row_idx].height = 15


def create_production_schedule_header(ws, last_filled_col: int) -> None:
    """Create the fixed header area for every Production Schedule sheet."""
    thin_bottom = Side(style="thin", color="000000")
    bottom_border = Border(bottom=thin_bottom)


    # Left side labels
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 16)
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 0, 28)

    for row_idx, label in [(4, "ZONE:"), (5, "FABRICATOR :"), (6, "FAB TYPE :")]:
        cell = ws.cell(row_idx, 1)
        cell.value = label
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.font = Font(name="Arial", size=10)

        value_cell = ws.cell(row_idx, 2)
        value_cell.font = Font(name="Arial", size=12, bold=True)
        value_cell.border = bottom_border
        value_cell.alignment = Alignment(horizontal="left", vertical="center")

    # Right side document information block
    # Example: if last column is O, labels go to L and merged fields go to M:O.
    # Keep the right block away from the generated logo area when the schedule is narrow.
    label_col = max(7, last_filled_col - 3)
    value_start_col = label_col + 1
    value_end_col = max(value_start_col, last_filled_col)

    right_labels = [
        (1, "Schedule No:"),
        (2, "Date Issued:"),
        (3, "Date Required:"),
        (4, "Project:"),
    ]

    for row_idx, label in right_labels:
        label_cell = ws.cell(row_idx, label_col)
        label_cell.value = label
        label_cell.alignment = Alignment(horizontal="right", vertical="center")
        label_cell.font = Font(name="Arial", size=10)

        merge_range = (
            f"{get_column_letter(value_start_col)}{row_idx}:"
            f"{get_column_letter(value_end_col)}{row_idx}"
        )
        try:
            ws.merge_cells(merge_range)
        except ValueError:
            pass

        # Apply bottom border to all cells in the merged range, not only top-left.
        for col_idx in range(value_start_col, value_end_col + 1):
            c = ws.cell(row_idx, col_idx)
            c.border = bottom_border
            c.font = Font(name="Arial", size=12, bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")

    apply_schedule_row_heights(ws)

    # Generated company logo in cells A1:F3, no image insertion.
    try_insert_logo(ws)

def apply_number_format(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = NUMBER_FORMAT


def apply_autofilter(ws, header_row: int = 1) -> None:
    if ws.max_row >= header_row and ws.max_column >= 1:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"


def write_matrix(ws, rows: List[List[Any]]) -> None:
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, value)


# =========================================================
# READ BOM AND CREATE COMBINED BOM
# =========================================================

def read_bom_workbook(bom_path: str) -> Tuple[List[dict], Dict[str, str], Dict[str, Any], List[List[Any]]]:
    """
    Returns:
      bom_rows: list of rows {sheet, row, code, description, qty}
      descriptions: code -> description
      unit_qty: sheet_name -> E2
      warnings: validation warnings
    """
    wb = load_workbook(bom_path, data_only=True)

    bom_rows: List[dict] = []
    descriptions: Dict[str, str] = {}
    unit_qty: Dict[str, Any] = {}
    warnings: List[List[Any]] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        unit_qty[sheet_name] = ws[BOM_UNIT_QTY_CELL].value
        seen_in_sheet: Dict[str, int] = {}

        for row_idx in range(BOM_START_ROW, ws.max_row + 1):
            raw_code = ws.cell(row_idx, BOM_CODE_COL).value
            code = clean_code(raw_code)
            description = ws.cell(row_idx, BOM_DESCRIPTION_COL).value
            raw_qty = ws.cell(row_idx, BOM_QTY_COL).value
            qty = to_number(raw_qty)

            if code is None:
                if description not in (None, "") or raw_qty not in (None, ""):
                    warnings.append([sheet_name, row_idx, "EMPTY_CODE_WITH_DATA", raw_code, description, raw_qty])
                continue

            if str(raw_code).strip().upper() != code:
                warnings.append([sheet_name, row_idx, "CODE_NORMALIZED", raw_code, code, raw_qty])

            if qty == 0 and raw_qty not in (None, "", 0, "0", "-"):
                warnings.append([sheet_name, row_idx, "QTY_NOT_NUMERIC", raw_code, description, raw_qty])

            seen_in_sheet[code] = seen_in_sheet.get(code, 0) + 1

            if code not in descriptions:
                descriptions[code] = "" if description is None else str(description).strip()

            bom_rows.append({
                "sheet": sheet_name,
                "row": row_idx,
                "code": code,
                "description": descriptions.get(code, ""),
                "qty": qty,
            })

        for code, count in seen_in_sheet.items():
            if count > 1:
                warnings.append([sheet_name, "", "DUPLICATE_CODE_IN_BOM_SHEET", code, "count", count])

    return bom_rows, descriptions, unit_qty, warnings


def build_summary(bom_rows: List[dict]) -> Dict[str, float]:
    summary: Dict[str, float] = {}
    for row in bom_rows:
        summary[row["code"]] = summary.get(row["code"], 0) + row["qty"]
    return summary


def build_unit_matrix(
    bom_rows: List[dict],
    descriptions: Dict[str, str],
    unit_qty: Dict[str, Any],
) -> List[List[Any]]:
    codes = sorted(descriptions.keys())
    units = list(unit_qty.keys())

    qty_by_code_unit: Dict[Tuple[str, str], float] = {}
    for row in bom_rows:
        key = (row["code"], row["sheet"])
        qty_by_code_unit[key] = qty_by_code_unit.get(key, 0) + row["qty"]

    rows: List[List[Any]] = []
    rows.append(["", "", "UNITS QTY"] + [unit_qty.get(unit) for unit in units])
    rows.append(["REF NUMBER", "DESCRIPTION", "TOTAL"] + units)

    for code in codes:
        qtys = [qty_by_code_unit.get((code, unit), 0) for unit in units]
        rows.append([code, descriptions.get(code, ""), sum(qtys)] + qtys)

    return rows


def build_parts_matrix(unit_matrix: List[List[Any]]) -> List[List[Any]]:
    if not unit_matrix:
        return []
    return list(map(list, zip(*unit_matrix)))


def create_combined_workbook(bom_path: str, combined_output_path: str) -> Tuple[Dict[str, float], Dict[str, str], List[List[Any]]]:
    bom_rows, descriptions, unit_qty, warnings = read_bom_workbook(bom_path)
    summary = build_summary(bom_rows)

    units_matrix = build_unit_matrix(bom_rows, descriptions, unit_qty)
    parts_matrix = build_parts_matrix(units_matrix)

    wb = Workbook()

    # Units
    ws_units = wb.active
    ws_units.title = "Units"
    write_matrix(ws_units, units_matrix)
    ws_units.freeze_panes = "D3"
    style_header_row(ws_units, 2)
    ws_units["C1"].font = Font(bold=True)
    apply_autofilter(ws_units, 2)
    apply_number_format(ws_units)
    auto_width(ws_units)

    # Parts
    ws_parts = wb.create_sheet("Parts")
    write_matrix(ws_parts, parts_matrix)
    ws_parts.freeze_panes = "C3"
    style_header_row(ws_parts, 1)
    apply_autofilter(ws_parts, 1)
    apply_number_format(ws_parts)
    auto_width(ws_parts)

    # Summary All
    ws_summary = wb.create_sheet("Summary All")
    rows = [["REF NUMBER", "DESCRIPTION", "TOTAL QTY"]]
    for code in sorted(summary.keys()):
        rows.append([code, descriptions.get(code, ""), summary[code]])
    write_matrix(ws_summary, rows)
    ws_summary.freeze_panes = "A2"
    style_header_row(ws_summary, 1)
    apply_autofilter(ws_summary, 1)
    apply_number_format(ws_summary)
    auto_width(ws_summary)

    # Checks
    ws_checks = wb.create_sheet("Checks")
    check_rows = [["Sheet", "Row", "Issue", "Value 1", "Value 2", "Value 3"]]
    check_rows.extend(warnings)
    if len(check_rows) == 1:
        check_rows.append(["OK", "", "No BOM warnings found", "", "", ""])
    write_matrix(ws_checks, check_rows)
    ws_checks.freeze_panes = "A2"
    style_header_row(ws_checks, 1, fill_color="FFF2CC")
    apply_autofilter(ws_checks, 1)
    auto_width(ws_checks)

    wb.save(combined_output_path)
    return summary, descriptions, warnings


# =========================================================
# LOG FILE AND PRODUCTION SCHEDULE
# =========================================================

def get_log_sheet_names(log_path: str) -> List[str]:
    wb = load_workbook(log_path, read_only=True, data_only=True)
    return wb.sheetnames


def get_used_log_columns(ws) -> List[int]:
    """
    Возвращает только нужные колонки LOG-листа.
    Пустые колонки перед QTY не попадают в Production Schedule.
    """
    valid_columns: List[int] = []

    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(LOG_HEADER_ROW, col_idx).value
        header_has_value = header is not None and str(header).strip() != ""

        data_has_value = False
        if not header_has_value:
            for row_idx in range(LOG_START_ROW, ws.max_row + 1):
                value = ws.cell(row_idx, col_idx).value
                if value is not None and str(value).strip() != "":
                    data_has_value = True
                    break

        if header_has_value or data_has_value:
            valid_columns.append(col_idx)

    return valid_columns


def read_log_sheet(log_path: str, sheet_name: str) -> Tuple[List[List[Any]], List[str], Dict[str, int], List[List[Any]], List[int]]:
    """
    Читает выбранный лист LOG file.

    Возвращает:
      rows: строки LOG без пустых колонок + QTY placeholder
      headers: заголовки без пустых колонок + QTY
      code_to_index: code -> index в rows
      warnings: предупреждения
      valid_columns: номера колонок из исходного LOG, которые реально используются
    """
    wb = load_workbook(log_path, data_only=False)
    ws = wb[sheet_name]

    valid_columns = get_used_log_columns(ws)

    headers: List[str] = []
    for col_idx in valid_columns:
        header = ws.cell(LOG_HEADER_ROW, col_idx).value
        if header is None or str(header).strip() == "":
            header = f"Column {col_idx}"
        headers.append(str(header).strip())

    headers.append("QTY")

    rows: List[List[Any]] = []
    code_to_index: Dict[str, int] = {}
    warnings: List[List[Any]] = []

    for row_idx in range(LOG_START_ROW, ws.max_row + 1):
        raw_code = ws.cell(row_idx, LOG_CODE_COL).value
        code = clean_code(raw_code)
        if code is None:
            continue

        row_values = [ws.cell(row_idx, col_idx).value for col_idx in valid_columns]

        # В списке valid_columns должна быть LOG_CODE_COL. Если нет — это ошибка формата.
        try:
            code_position = valid_columns.index(LOG_CODE_COL)
            row_values[code_position] = code
        except ValueError:
            warnings.append([sheet_name, row_idx, "LOG_CODE_COLUMN_NOT_IN_VALID_COLUMNS", raw_code, "", ""])

        row_values.append(0)  # QTY placeholder

        if code in code_to_index:
            warnings.append([sheet_name, row_idx, "DUPLICATE_CODE_IN_LOG", code, "", ""])
        else:
            code_to_index[code] = len(rows)

        if raw_code is not None and str(raw_code).strip().upper() != code:
            warnings.append([sheet_name, row_idx, "LOG_CODE_NORMALIZED", raw_code, code, ""])

        rows.append(row_values)

    return rows, headers, code_to_index, warnings, valid_columns


def create_schedule_sheet_from_log(
    out_wb: Workbook,
    log_path: str,
    log_sheet_name: str,
    summary: Dict[str, float],
    descriptions: Dict[str, str],
    existing_titles: set,
) -> Dict[str, Any]:
    """Create one Production Schedule sheet for one selected LOG sheet."""
    log_rows, headers, code_to_index, log_warnings, valid_columns = read_log_sheet(log_path, log_sheet_name)
    log_codes = set(code_to_index.keys())

    total_matched_qty = 0
    matched_codes: List[str] = []

    # Fill QTY using BOM Summary codes
    for code, qty in summary.items():
        if code in code_to_index:
            row_index = code_to_index[code]
            log_rows[row_index][-1] = qty
            total_matched_qty += qty
            matched_codes.append(code)

    # Keep only schedule rows where QTY > 0
    schedule_rows = [row for row in log_rows if to_number(row[-1]) != 0]

    sheet_title = safe_sheet_title(log_sheet_name, existing_titles)
    ws = out_wb.create_sheet(sheet_title)

    max_columns = len(headers)

    # Create fixed header area in rows 1-6.
    create_production_schedule_header(ws, max_columns)

    # Write table headers to row 7
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(SCHEDULE_HEADER_ROW, col_idx)
        cell.value = header

    # QTY is always the last column
    qty_col = max_columns
    ws.cell(SCHEDULE_HEADER_ROW, qty_col).value = "QTY"

    # Apply standard header style
    style_header_row(ws, SCHEDULE_HEADER_ROW)

    # Write schedule data from row 8
    for row_idx, row in enumerate(schedule_rows, start=SCHEDULE_DATA_ROW):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row_idx, col_idx, value)

    # Center alignment начиная с 3-й колонки, начиная с 7-й строки
    for row_idx in range(SCHEDULE_HEADER_ROW, ws.max_row + 1):
        for col_idx in range(3, ws.max_column + 1):
            ws.cell(row_idx, col_idx).alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

    # Formatting
    ws.freeze_panes = f"A{SCHEDULE_DATA_ROW}"
    apply_autofilter(ws, SCHEDULE_HEADER_ROW)
    apply_number_format(ws)
    auto_width(ws)
    ws.column_dimensions[get_column_letter(qty_col)].width = 12

    # SUBTOTAL row for QTY
    if schedule_rows:
        total_row = SCHEDULE_DATA_ROW + len(schedule_rows)
        ws.cell(total_row, qty_col - 1, "TOTAL")
        ws.cell(total_row, qty_col, f"=SUBTOTAL(9,{get_column_letter(qty_col)}{SCHEDULE_DATA_ROW}:{get_column_letter(qty_col)}{total_row - 1})")
        for col_idx in range(1, max_columns + 1):
            c = ws.cell(total_row, col_idx)
            c.font = Font(bold=True)
            c.border = Border(top=Side(style="thin"), bottom=Side(style="double"))

    # Row heights after all schedule rows and total row are created
    apply_schedule_row_heights(ws)

    # Checks for the report
    missing_from_this_log = sorted(code for code in summary.keys() if code not in log_codes)
    unused_in_this_log = sorted(code for code in log_codes if code not in summary)
    scheduled_codes = sorted(matched_codes)

    return {
        "log_sheet": log_sheet_name,
        "schedule_sheet": sheet_title,
        "matched_codes_count": len(matched_codes),
        "schedule_rows_count": len(schedule_rows),
        "total_matched_qty": total_matched_qty,
        "missing_from_this_log": missing_from_this_log,
        "unused_in_this_log": unused_in_this_log,
        "log_warnings": log_warnings,
        "valid_log_columns": valid_columns,
        "scheduled_codes": scheduled_codes,
    }

def create_drawing_codes_sheet(wb: Workbook, reports: List[Dict[str, Any]], summary: Dict[str, float], descriptions: Dict[str, str]) -> None:
    """Create a Drawing Codes sheet for all scheduled part codes."""
    ws = wb.create_sheet("Drawing Codes")

    drawing_map: Dict[str, Dict[str, Any]] = {}
    for rep in reports:
        log_sheet = rep.get("log_sheet", "")
        for part_code in rep.get("scheduled_codes", []):
            code = clean_code(part_code)
            drawing_code = get_drawing_code(code)
            if code is None or drawing_code is None:
                continue

            if drawing_code not in drawing_map:
                drawing_map[drawing_code] = {
                    "parts": set(),
                    "log_sheets": set(),
                    "total_qty": 0,
                }

            drawing_map[drawing_code]["parts"].add(code)
            drawing_map[drawing_code]["log_sheets"].add(log_sheet)
            drawing_map[drawing_code]["total_qty"] += to_number(summary.get(code, 0))

    rows = [["DRAWING CODE", "PART CODES", "PART COUNT", "TOTAL QTY", "LOG SHEETS"]]
    for drawing_code in sorted(drawing_map.keys()):
        item = drawing_map[drawing_code]
        parts = sorted(item["parts"])
        log_sheets = sorted(item["log_sheets"])
        rows.append([
            drawing_code,
            ", ".join(parts),
            len(parts),
            item["total_qty"],
            ", ".join(log_sheets),
        ])

    if len(rows) == 1:
        rows.append(["OK", "No scheduled part codes found", 0, 0, ""])

    write_matrix(ws, rows)
    style_header_row(ws, 1, fill_color="D9EAD3")
    ws.freeze_panes = "A2"
    apply_autofilter(ws, 1)
    apply_number_format(ws)
    auto_width(ws, max_width=80)


def create_production_schedule_workbook(
    log_path: str,
    selected_log_sheets: List[str],
    summary: Dict[str, float],
    descriptions: Dict[str, str],
    schedule_output_path: str,
) -> None:
    wb = Workbook()
    report_ws = wb.active
    report_ws.title = "Report"
    existing_titles = {"Report"}

    reports: List[Dict[str, Any]] = []
    for sheet_name in selected_log_sheets:
        reports.append(
            create_schedule_sheet_from_log(
                out_wb=wb,
                log_path=log_path,
                log_sheet_name=sheet_name,
                summary=summary,
                descriptions=descriptions,
                existing_titles=existing_titles,
            )
        )

    # Drawing Codes
    create_drawing_codes_sheet(wb, reports, summary, descriptions)
    existing_titles.add("Drawing Codes")

    # Report
    report_rows = [
        ["Production Schedule Report"],
        [],
        ["Selected LOG sheet", "Schedule sheet", "Matched codes", "Schedule rows", "Total QTY", "Used LOG columns"],
    ]
    for rep in reports:
        report_rows.append([
            rep["log_sheet"],
            rep["schedule_sheet"],
            rep["matched_codes_count"],
            rep["schedule_rows_count"],
            rep["total_matched_qty"],
            ", ".join(str(c) for c in rep["valid_log_columns"]),
        ])
    write_matrix(report_ws, report_rows)
    report_ws["A1"].font = Font(bold=True, size=14)
    style_header_row(report_ws, 3)
    apply_number_format(report_ws)
    auto_width(report_ws)

    # BOM codes not found in any selected LOG sheet
    selected_log_codes = set()
    all_log_warnings: List[List[Any]] = []

    for rep in reports:
        all_log_warnings.extend(rep["log_warnings"])
        _, _, code_to_index, _, _ = read_log_sheet(log_path, rep["log_sheet"])
        selected_log_codes.update(code_to_index.keys())

    missing_all = sorted(code for code in summary.keys() if code not in selected_log_codes)
    ws_missing = wb.create_sheet("Missing in Selected LOG")
    rows = [["REF NUMBER", "DESCRIPTION", "TOTAL QTY", "Issue"]]
    for code in missing_all:
        rows.append([code, descriptions.get(code, ""), summary.get(code, 0), "Code exists in BOM but not in selected LOG sheets"])
    if len(rows) == 1:
        rows.append(["OK", "", "", "All BOM codes were found in selected LOG sheets"])
    write_matrix(ws_missing, rows)
    style_header_row(ws_missing, 1, fill_color="F4CCCC")
    ws_missing.freeze_panes = "A2"
    apply_autofilter(ws_missing, 1)
    apply_number_format(ws_missing)
    auto_width(ws_missing)

    # BOM Summary
    ws_summary = wb.create_sheet("BOM Summary")
    rows = [["REF NUMBER", "DESCRIPTION", "TOTAL QTY"]]
    for code in sorted(summary.keys()):
        rows.append([code, descriptions.get(code, ""), summary[code]])
    write_matrix(ws_summary, rows)
    style_header_row(ws_summary, 1)
    ws_summary.freeze_panes = "A2"
    apply_autofilter(ws_summary, 1)
    apply_number_format(ws_summary)
    auto_width(ws_summary)

    # LOG Checks
    ws_warn = wb.create_sheet("LOG Checks")
    rows = [["LOG Sheet", "Row", "Issue", "Value 1", "Value 2", "Value 3"]]
    rows.extend(all_log_warnings)
    if len(rows) == 1:
        rows.append(["OK", "", "No LOG warnings found", "", "", ""])
    write_matrix(ws_warn, rows)
    style_header_row(ws_warn, 1, fill_color="FFF2CC")
    ws_warn.freeze_panes = "A2"
    apply_autofilter(ws_warn, 1)
    auto_width(ws_warn)

    wb.save(schedule_output_path)


# =========================================================
# GUI
# =========================================================

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("BOM Production Automation v13")
        self.geometry("980x720")
        self.minsize(800, 600)
        self.resizable(True, True)

        self.bom_path_var = tk.StringVar()
        self.log_path_var = tk.StringVar()
        self.combined_output_var = tk.StringVar()
        self.schedule_output_var = tk.StringVar()
        self.log_sheets: List[str] = []
        self.sheet_checkboxes: Dict[str, ctk.CTkCheckBox] = {}
        self.sheet_vars: Dict[str, tk.BooleanVar] = {}

        self._build_ui()

    def _build_ui(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)

        # ── Main scroll container ──────────────────────────────
        main_frame = ctk.CTkScrollableFrame(self, corner_radius=0, fg_color="transparent")
        main_frame.grid(row=0, column=0, sticky="nsew", padx=0, pady=0)
        main_frame.grid_columnconfigure(1, weight=1)

        # ── Title ─────────────────────────────────────────────
        title = ctk.CTkLabel(
            main_frame,
            text="⚙️  BOM Production Automation",
            font=ctk.CTkFont(size=22, weight="bold"),
        )
        title.grid(row=0, column=0, columnspan=3, pady=(24, 4), padx=24, sticky="w")

        subtitle = ctk.CTkLabel(
            main_frame,
            text="Automates BOM → Combined BOM → Production Schedule workflow",
            font=ctk.CTkFont(size=13),
            text_color="gray",
        )
        subtitle.grid(row=1, column=0, columnspan=3, pady=(0, 20), padx=24, sticky="w")

        # ── Section: Input Files ───────────────────────────────
        self._section_label(main_frame, "📂  Input Files", row=2)

        self._file_row(main_frame, row=3,
                       label="BOM File (.xlsx / .xlsm)",
                       var=self.bom_path_var,
                       command=self.select_bom,
                       btn_text="Browse")

        self._file_row(main_frame, row=4,
                       label="LOG File (.xlsx / .xlsm)",
                       var=self.log_path_var,
                       command=self.select_log,
                       btn_text="Browse")

        # ── Section: LOG Sheets ────────────────────────────────
        self._section_label(main_frame, "📋  Select LOG Sheets", row=5)

        sheet_outer = ctk.CTkFrame(main_frame, fg_color="transparent")
        sheet_outer.grid(row=6, column=0, columnspan=3, sticky="ew", padx=24, pady=(0, 8))
        sheet_outer.grid_columnconfigure(0, weight=1)

        # scrollable checkbox area
        self.sheet_frame = ctk.CTkScrollableFrame(
            sheet_outer,
            height=160,
            label_text="",
            fg_color=("gray92", "gray17"),
            corner_radius=8,
        )
        self.sheet_frame.grid(row=0, column=0, sticky="ew")
        self.sheet_frame.grid_columnconfigure(0, weight=1)

        self.no_sheets_label = ctk.CTkLabel(
            self.sheet_frame,
            text="Load a LOG file to see available sheets",
            text_color="gray",
        )
        self.no_sheets_label.grid(row=0, column=0, pady=20)

        # Select all / Clear buttons
        btn_row = ctk.CTkFrame(sheet_outer, fg_color="transparent")
        btn_row.grid(row=1, column=0, sticky="w", pady=(6, 0))
        ctk.CTkButton(btn_row, text="Select All", width=110,
                      command=self.select_all_sheets).pack(side="left", padx=(0, 8))
        ctk.CTkButton(btn_row, text="Clear All", width=110, fg_color="gray40",
                      hover_color="gray30", command=self.clear_sheets).pack(side="left")

        # ── Section: Output Files ──────────────────────────────
        self._section_label(main_frame, "💾  Output Files", row=7)

        self._file_row(main_frame, row=8,
                       label="Combined BOM Output",
                       var=self.combined_output_var,
                       command=self.select_combined_output,
                       btn_text="Save as")

        self._file_row(main_frame, row=9,
                       label="Production Schedule Output",
                       var=self.schedule_output_var,
                       command=self.select_schedule_output,
                       btn_text="Save as")

        # ── Progress bar ───────────────────────────────────────
        self.progress = ctk.CTkProgressBar(main_frame, mode="indeterminate", height=6)
        self.progress.grid(row=10, column=0, columnspan=3, sticky="ew", padx=24, pady=(16, 0))
        self.progress.set(0)

        # ── RUN button ─────────────────────────────────────────
        self.run_button = ctk.CTkButton(
            main_frame,
            text="▶  RUN",
            font=ctk.CTkFont(size=16, weight="bold"),
            height=48,
            corner_radius=10,
            command=self.run,
        )
        self.run_button.grid(row=11, column=0, columnspan=3, sticky="ew", padx=24, pady=(12, 4))

        # ── Status label ───────────────────────────────────────
        self.status_label = ctk.CTkLabel(
            main_frame,
            text="● Ready",
            font=ctk.CTkFont(size=13),
            text_color="#4CAF50",
            anchor="w",
        )
        self.status_label.grid(row=12, column=0, columnspan=3, sticky="w", padx=26, pady=(2, 16))

    # ── Helpers ────────────────────────────────────────────────

    def _section_label(self, parent, text: str, row: int):
        frame = ctk.CTkFrame(parent, fg_color="transparent")
        frame.grid(row=row, column=0, columnspan=3, sticky="ew", padx=24, pady=(16, 4))
        frame.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(frame, text=text, font=ctk.CTkFont(size=14, weight="bold")).grid(row=0, column=0, sticky="w")
        sep = ctk.CTkFrame(frame, height=2, fg_color=("gray80", "gray30"), corner_radius=1)
        sep.grid(row=0, column=1, sticky="ew", padx=(12, 0), pady=6)

    def _file_row(self, parent, row: int, label: str, var: tk.StringVar, command, btn_text: str):
        ctk.CTkLabel(parent, text=label, font=ctk.CTkFont(size=13), anchor="w").grid(
            row=row, column=0, sticky="w", padx=(24, 8), pady=4)
        entry = ctk.CTkEntry(parent, textvariable=var, height=36, corner_radius=8)
        entry.grid(row=row, column=1, sticky="ew", padx=(0, 8), pady=4)
        ctk.CTkButton(parent, text=btn_text, width=100, height=36,
                      corner_radius=8, command=command).grid(row=row, column=2, padx=(0, 24), pady=4)

    # ── Status ─────────────────────────────────────────────────

    def set_status(self, text: str, color: str = "#4CAF50"):
        self.status_label.configure(text=f"● {text}", text_color=color)
        self.update_idletasks()

    # ── File selection ─────────────────────────────────────────

    def select_bom(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xlsm")])
        if not path:
            return
        self.bom_path_var.set(path)
        base = Path(path).with_suffix("")
        self.combined_output_var.set(str(base) + " - Combined BOM.xlsx")
        self.schedule_output_var.set(str(base) + " - Production Schedule.xlsx")

    def select_log(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xlsm")])
        if not path:
            return
        self.log_path_var.set(path)
        try:
            self.log_sheets = get_log_sheet_names(path)
            self._populate_sheets(self.log_sheets)
            self.set_status(f"Loaded {len(self.log_sheets)} LOG sheets")
        except Exception as exc:
            messagebox.showerror("LOG error", f"Cannot read LOG file:\n{exc}")
            self.set_status("Error reading LOG file", "#F44336")

    def _populate_sheets(self, sheets: List[str]):
        for widget in self.sheet_frame.winfo_children():
            widget.destroy()
        self.sheet_checkboxes.clear()
        self.sheet_vars.clear()

        if not sheets:
            ctk.CTkLabel(self.sheet_frame, text="No sheets found", text_color="gray").grid(row=0, column=0, pady=20)
            return

        for i, sheet in enumerate(sheets):
            var = tk.BooleanVar(value=True)
            cb = ctk.CTkCheckBox(self.sheet_frame, text=sheet, variable=var, font=ctk.CTkFont(size=13))
            cb.grid(row=i, column=0, sticky="w", padx=12, pady=3)
            self.sheet_checkboxes[sheet] = cb
            self.sheet_vars[sheet] = var

    def select_all_sheets(self):
        for var in self.sheet_vars.values():
            var.set(True)

    def clear_sheets(self):
        for var in self.sheet_vars.values():
            var.set(False)

    def select_combined_output(self):
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if path:
            self.combined_output_var.set(path)

    def select_schedule_output(self):
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if path:
            self.schedule_output_var.set(path)

    def get_selected_sheets(self) -> List[str]:
        return [sheet for sheet, var in self.sheet_vars.items() if var.get()]

    # ── Validation ─────────────────────────────────────────────

    def validate_inputs(self) -> bool:
        if not self.bom_path_var.get():
            messagebox.showerror("Missing BOM", "Please select a BOM file.")
            return False
        if not self.log_path_var.get():
            messagebox.showerror("Missing LOG", "Please select a LOG file.")
            return False
        if not self.get_selected_sheets():
            messagebox.showerror("Missing LOG sheets", "Please select at least one LOG sheet.")
            return False
        if not self.combined_output_var.get():
            messagebox.showerror("Missing output", "Please select Combined BOM output path.")
            return False
        if not self.schedule_output_var.get():
            messagebox.showerror("Missing output", "Please select Production Schedule output path.")
            return False
        return True

    # ── RUN ────────────────────────────────────────────────────

    def run(self):
        if not self.validate_inputs():
            return

        bom_path = self.bom_path_var.get()
        log_path = self.log_path_var.get()
        selected_sheets = self.get_selected_sheets()
        combined_output = self.combined_output_var.get()
        schedule_output = self.schedule_output_var.get()

        try:
            self.run_button.configure(state="disabled", text="⏳  Processing...")
            self.progress.start()

            self.set_status("Creating Combined BOM...", "#FFC107")
            summary, descriptions, _ = create_combined_workbook(bom_path, combined_output)

            self.set_status("Creating Production Schedule...", "#FFC107")
            create_production_schedule_workbook(
                log_path=log_path,
                selected_log_sheets=selected_sheets,
                summary=summary,
                descriptions=descriptions,
                schedule_output_path=schedule_output,
            )

            self.progress.stop()
            self.progress.set(1)
            self.set_status("Done — files created successfully!", "#4CAF50")
            messagebox.showinfo(
                "Success ✅",
                f"Files created successfully!\n\n"
                f"📊 Combined BOM:\n{combined_output}\n\n"
                f"📋 Production Schedule:\n{schedule_output}\n\n"
                f"✔ Selected LOG sheets: {len(selected_sheets)}",
            )
        except Exception as exc:
            self.progress.stop()
            self.progress.set(0)
            self.set_status(f"Error: {exc}", "#F44336")
            messagebox.showerror("Error ❌", str(exc))
        finally:
            self.run_button.configure(state="normal", text="▶  RUN")


def main():
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
