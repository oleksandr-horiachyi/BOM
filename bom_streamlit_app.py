"""
BOM Production Automation — Streamlit Web Version
FabTools | https://oleksandr-horiachyi.github.io/BOM/
"""

import io
import re
import zipfile
from typing import Any, Dict, List, Optional, Tuple

import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="BOM Production Automation — FabTools",
    page_icon="⚙️",
    layout="centered",
)

# ─────────────────────────────────────────────────────────────
# CUSTOM CSS  (Concept B — Clean Corporate style)
# ─────────────────────────────────────────────────────────────

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Source+Sans+3:wght@400;500;600;700&display=swap');

html, body, [class*="css"] {
    font-family: 'Source Sans 3', sans-serif;
}

/* Hero banner */
.hero-banner {
    background: linear-gradient(135deg, #0C447C 0%, #185FA5 55%, #378ADD 100%);
    border-radius: 12px;
    padding: 36px 36px 28px;
    color: white;
    margin-bottom: 32px;
}
.hero-banner h1 {
    font-size: 28px;
    font-weight: 700;
    margin: 0 0 8px 0;
    line-height: 1.2;
}
.hero-banner p {
    font-size: 15px;
    color: rgba(255,255,255,0.82);
    margin: 0 0 20px 0;
    line-height: 1.6;
}
.hero-steps {
    display: flex;
    align-items: center;
    gap: 10px;
    flex-wrap: wrap;
}
.hero-step {
    display: flex;
    align-items: center;
    gap: 8px;
    font-size: 14px;
    font-weight: 600;
    color: rgba(255,255,255,0.95);
}
.step-num {
    width: 24px; height: 24px;
    border-radius: 50%;
    background: rgba(255,255,255,0.2);
    border: 1.5px solid rgba(255,255,255,0.4);
    display: flex; align-items: center; justify-content: center;
    font-size: 12px; font-weight: 700;
    text-align: center; line-height: 24px;
}
.hero-arrow { color: rgba(255,255,255,0.4); font-size: 16px; }

/* Trust badge */
.trust-badge {
    display: inline-flex;
    align-items: center;
    gap: 8px;
    background: rgba(255,255,255,0.1);
    border: 1px solid rgba(255,255,255,0.2);
    border-radius: 8px;
    padding: 7px 14px;
    font-size: 13px;
    color: rgba(255,255,255,0.85);
    margin-top: 16px;
}

/* Section headers */
.section-header {
    font-size: 13px;
    font-weight: 600;
    letter-spacing: 1.2px;
    text-transform: uppercase;
    color: #185FA5;
    margin: 0 0 12px 0;
    padding-top: 8px;
    border-top: 2px solid #E6F1FB;
}

/* Info box */
.info-box {
    background: #E6F1FB;
    border-left: 3px solid #185FA5;
    border-radius: 0 8px 8px 0;
    padding: 10px 14px;
    font-size: 13px;
    color: #0C447C;
    line-height: 1.55;
    margin-bottom: 12px;
}

/* Output box */
.output-box {
    background: #F0F9F5;
    border: 1px solid #A8D5BF;
    border-radius: 8px;
    padding: 14px 18px;
    margin-bottom: 10px;
}
.output-box-title {
    font-size: 15px;
    font-weight: 700;
    color: #0A5C3A;
    margin-bottom: 4px;
}
.output-box-sheets {
    font-size: 12px;
    color: #3D8A6B;
}
.output-box-tag {
    display: inline-block;
    background: #C6EBD9;
    color: #0A5C3A;
    border-radius: 3px;
    padding: 1px 7px;
    margin: 2px 2px 0 0;
    font-size: 11px;
}

/* Footer */
.footer-bar {
    text-align: center;
    font-size: 12px;
    color: #888;
    margin-top: 40px;
    padding-top: 20px;
    border-top: 1px solid #EFF1F5;
}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────
# SETTINGS  (same as v13)
# ─────────────────────────────────────────────────────────────

BOM_START_ROW       = 3
BOM_CODE_COL        = 1
BOM_DESCRIPTION_COL = 2
BOM_QTY_COL         = 5
BOM_UNIT_QTY_CELL   = "E2"

LOG_HEADER_ROW = 3
LOG_START_ROW  = 4
LOG_CODE_COL   = 1

SCHEDULE_HEADER_ROW = 7
SCHEDULE_DATA_ROW   = 8

NUMBER_FORMAT = '# ##0;-# ##0;-'

# ─────────────────────────────────────────────────────────────
# CORE FUNCTIONS  (identical logic to v13, no GUI dependencies)
# ─────────────────────────────────────────────────────────────

def clean_code(value: Any) -> Optional[str]:
    if value is None:
        return None
    # Fix: Excel sometimes stores whole numbers as float (4003 → 4003.0)
    # Convert to int first so str() gives "4003" not "4003.0"
    if isinstance(value, float) and value == int(value):
        value = int(value)
    text = str(value).replace("\u00a0", " ").strip().upper()
    text = re.sub(r"\s+", " ", text)
    if text in ("", "NONE", "NAN"):
        return None
    return text

def to_number(value: Any) -> float:
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return value
    text = str(value).replace(",", ".").strip()
    if text in ("", "-"):
        return 0
    try:
        return float(text)
    except ValueError:
        return 0

def get_drawing_code(part_code: Any) -> Optional[str]:
    code = clean_code(part_code)
    if code is None:
        return None
    return code[:6]

def safe_sheet_title(title: str, existing: Optional[set] = None) -> str:
    if existing is None:
        existing = set()
    title = re.sub(r"[\\/*?:\[\]]", "-", str(title)).strip()
    title = title[:31] if title else "Sheet"
    base = title
    counter = 1
    while title in existing:
        suffix = f"_{counter}"
        title = base[:31 - len(suffix)] + suffix
        counter += 1
    existing.add(title)
    return title

def auto_width(ws, min_width: int = 8, max_width: int = 45) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = max(min_width, min(max_width, max_len + 2))

def style_header_row(ws, row: int = 1, fill_color: str = "D9EAF7") -> None:
    fill = PatternFill("solid", fgColor=fill_color)
    thin = Side(style="thin", color="B7B7B7")
    for cell in ws[row]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)

def apply_number_format(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = NUMBER_FORMAT

def apply_autofilter(ws, header_row: int = 1) -> None:
    if ws.max_row >= header_row and ws.max_column >= 1:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"

def write_matrix(ws, rows: List[List[Any]]) -> None:
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, value)

def apply_schedule_row_heights(ws) -> None:
    for row_idx in range(1, 7):
        ws.row_dimensions[row_idx].height = 20
    last_row = max(ws.max_row, SCHEDULE_DATA_ROW)
    for row_idx in range(7, last_row + 1):
        ws.row_dimensions[row_idx].height = 15

def create_production_schedule_header(ws, last_filled_col: int) -> None:
    thin_bottom = Side(style="thin", color="000000")
    bottom_border = Border(bottom=thin_bottom)

    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 16)
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 0, 28)

    for row_idx, label in [(4, "ZONE:"), (5, "FABRICATOR :"), (6, "FAB TYPE :")]:
        cell = ws.cell(row_idx, 1)
        cell.value = label
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.font = Font(name="Arial", size=10)
        value_cell = ws.cell(row_idx, 2)
        value_cell.font = Font(name="Arial", size=12, bold=True)
        value_cell.border = bottom_border
        value_cell.alignment = Alignment(horizontal="left", vertical="center")

    label_col = max(7, last_filled_col - 3)
    value_start_col = label_col + 1
    value_end_col = max(value_start_col, last_filled_col)

    for row_idx, label in [(1, "Schedule No:"), (2, "Date Issued:"), (3, "Date Required:"), (4, "Project:")]:
        label_cell = ws.cell(row_idx, label_col)
        label_cell.value = label
        label_cell.alignment = Alignment(horizontal="right", vertical="center")
        label_cell.font = Font(name="Arial", size=10)
        merge_range = (
            f"{get_column_letter(value_start_col)}{row_idx}:"
            f"{get_column_letter(value_end_col)}{row_idx}"
        )
        try:
            ws.merge_cells(merge_range)
        except ValueError:
            pass
        for col_idx in range(value_start_col, value_end_col + 1):
            c = ws.cell(row_idx, col_idx)
            c.border = bottom_border
            c.font = Font(name="Arial", size=12, bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")

    apply_schedule_row_heights(ws)
    # logo is inserted separately when logo_bytes is available

def try_insert_logo(ws, logo_bytes: Optional[bytes]) -> None:
    """Insert logo into cell A1 of the worksheet, scaled to fit rows 1–3."""
    if not logo_bytes:
        return
    try:
        from PIL import Image as PILImage
        from openpyxl.drawing.image import Image as XLImage

        # Get dimensions with PIL (no temp file needed)
        pil_img = PILImage.open(io.BytesIO(logo_bytes))
        original_width, original_height = pil_img.size

        # Scale to fit rows 1-3 (each row = 20pt, 1pt ≈ 1.333px)
        target_height = int(20 * 3 * 1.333)
        scale = target_height / original_height
        target_width = int(original_width * scale)

        # Pass BytesIO directly — no temp file, no deletion issues
        xl_img = XLImage(io.BytesIO(logo_bytes))
        xl_img.width = target_width
        xl_img.height = target_height
        ws.add_image(xl_img, "A1")
    except Exception as e:
        pass  # Logo insert is optional — never break the report

# ─── BOM ────────────────────────────────────────────────────

def read_bom_workbook(bom_bytes: bytes):
    wb = load_workbook(io.BytesIO(bom_bytes), data_only=True)
    bom_rows, descriptions, unit_qty, warnings = [], {}, {}, []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        unit_qty[sheet_name] = ws[BOM_UNIT_QTY_CELL].value
        seen_in_sheet: Dict[str, int] = {}

        for row_idx in range(BOM_START_ROW, ws.max_row + 1):
            raw_code = ws.cell(row_idx, BOM_CODE_COL).value
            code = clean_code(raw_code)
            description = ws.cell(row_idx, BOM_DESCRIPTION_COL).value
            raw_qty = ws.cell(row_idx, BOM_QTY_COL).value
            qty = to_number(raw_qty)

            if code is None:
                continue

            seen_in_sheet[code] = seen_in_sheet.get(code, 0) + 1
            if code not in descriptions:
                descriptions[code] = "" if description is None else str(description).strip()

            bom_rows.append({"sheet": sheet_name, "row": row_idx, "code": code,
                             "description": descriptions.get(code, ""), "qty": qty})

        for code, count in seen_in_sheet.items():
            if count > 1:
                warnings.append([sheet_name, "", "DUPLICATE_CODE_IN_BOM_SHEET", code, "count", count])

    return bom_rows, descriptions, unit_qty, warnings

def build_summary(bom_rows):
    summary = {}
    for row in bom_rows:
        summary[row["code"]] = summary.get(row["code"], 0) + row["qty"]
    return summary

def build_unit_matrix(bom_rows, descriptions, unit_qty):
    codes = sorted(descriptions.keys())
    units = list(unit_qty.keys())
    qty_by_code_unit = {}
    for row in bom_rows:
        key = (row["code"], row["sheet"])
        qty_by_code_unit[key] = qty_by_code_unit.get(key, 0) + row["qty"]

    rows = [["", "", "UNITS QTY"] + [unit_qty.get(u) for u in units]]
    rows.append(["REF NUMBER", "DESCRIPTION", "TOTAL"] + units)
    for code in codes:
        qtys = [qty_by_code_unit.get((code, u), 0) for u in units]
        rows.append([code, descriptions.get(code, ""), sum(qtys)] + qtys)
    return rows

def create_combined_workbook_bytes(bom_bytes: bytes) -> Tuple[bytes, Dict, Dict, List]:
    bom_rows, descriptions, unit_qty, warnings = read_bom_workbook(bom_bytes)
    summary = build_summary(bom_rows)
    units_matrix = build_unit_matrix(bom_rows, descriptions, unit_qty)
    parts_matrix = list(map(list, zip(*units_matrix))) if units_matrix else []

    wb = Workbook()

    ws_units = wb.active
    ws_units.title = "Units"
    write_matrix(ws_units, units_matrix)
    ws_units.freeze_panes = "D3"
    style_header_row(ws_units, 2)
    ws_units["C1"].font = Font(bold=True)
    apply_autofilter(ws_units, 2)
    apply_number_format(ws_units)
    auto_width(ws_units)

    ws_parts = wb.create_sheet("Parts")
    write_matrix(ws_parts, parts_matrix)
    ws_parts.freeze_panes = "C3"
    style_header_row(ws_parts, 1)
    apply_autofilter(ws_parts, 1)
    apply_number_format(ws_parts)
    auto_width(ws_parts)

    ws_summary = wb.create_sheet("Summary All")
    rows = [["REF NUMBER", "DESCRIPTION", "TOTAL QTY"]]
    for code in sorted(summary.keys()):
        rows.append([code, descriptions.get(code, ""), summary[code]])
    write_matrix(ws_summary, rows)
    ws_summary.freeze_panes = "A2"
    style_header_row(ws_summary, 1)
    apply_autofilter(ws_summary, 1)
    apply_number_format(ws_summary)
    auto_width(ws_summary)

    ws_checks = wb.create_sheet("Checks")
    check_rows = [["Sheet", "Row", "Issue", "Value 1", "Value 2", "Value 3"]]
    check_rows.extend(warnings)
    if len(check_rows) == 1:
        check_rows.append(["OK", "", "No BOM warnings found", "", "", ""])
    write_matrix(ws_checks, check_rows)
    ws_checks.freeze_panes = "A2"
    style_header_row(ws_checks, 1, fill_color="FFF2CC")
    apply_autofilter(ws_checks, 1)
    auto_width(ws_checks)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), summary, descriptions, warnings

# ─── LOG ────────────────────────────────────────────────────

def get_log_sheet_names(log_bytes: bytes) -> List[str]:
    wb = load_workbook(io.BytesIO(log_bytes), read_only=True, data_only=True)
    return wb.sheetnames

def get_used_log_columns(ws) -> List[int]:
    valid_columns = []
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(LOG_HEADER_ROW, col_idx).value
        header_has_value = header is not None and str(header).strip() != ""
        data_has_value = False
        if not header_has_value:
            for row_idx in range(LOG_START_ROW, ws.max_row + 1):
                value = ws.cell(row_idx, col_idx).value
                if value is not None and str(value).strip() != "":
                    data_has_value = True
                    break
        if header_has_value or data_has_value:
            valid_columns.append(col_idx)
    return valid_columns

def read_log_sheet(log_bytes: bytes, sheet_name: str):
    wb = load_workbook(io.BytesIO(log_bytes), data_only=False)
    ws = wb[sheet_name]
    valid_columns = get_used_log_columns(ws)

    headers = []
    for col_idx in valid_columns:
        header = ws.cell(LOG_HEADER_ROW, col_idx).value
        if header is None or str(header).strip() == "":
            header = f"Column {col_idx}"
        headers.append(str(header).strip())
    headers.append("QTY")

    rows, code_to_index, warnings = [], {}, []
    for row_idx in range(LOG_START_ROW, ws.max_row + 1):
        raw_code = ws.cell(row_idx, LOG_CODE_COL).value
        code = clean_code(raw_code)
        if code is None:
            continue
        row_values = [ws.cell(row_idx, col_idx).value for col_idx in valid_columns]
        try:
            code_position = valid_columns.index(LOG_CODE_COL)
            row_values[code_position] = code
        except ValueError:
            pass
        row_values.append(0)
        if code not in code_to_index:
            code_to_index[code] = len(rows)
        rows.append(row_values)

    return rows, headers, code_to_index, warnings, valid_columns

def create_schedule_sheet_from_log(out_wb, log_bytes, log_sheet_name, summary, descriptions, existing_titles, logo_bytes=None):
    log_rows, headers, code_to_index, log_warnings, valid_columns = read_log_sheet(log_bytes, log_sheet_name)
    log_codes = set(code_to_index.keys())
    total_matched_qty = 0
    matched_codes = []

    for code, qty in summary.items():
        if code in code_to_index:
            log_rows[code_to_index[code]][-1] = qty
            total_matched_qty += qty
            matched_codes.append(code)

    schedule_rows = [row for row in log_rows if to_number(row[-1]) != 0]
    sheet_title = safe_sheet_title(log_sheet_name, existing_titles)
    ws = out_wb.create_sheet(sheet_title)
    max_columns = len(headers)

    create_production_schedule_header(ws, max_columns)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(SCHEDULE_HEADER_ROW, col_idx)
        cell.value = header
    ws.cell(SCHEDULE_HEADER_ROW, max_columns).value = "QTY"
    style_header_row(ws, SCHEDULE_HEADER_ROW)

    for row_idx, row in enumerate(schedule_rows, start=SCHEDULE_DATA_ROW):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row_idx, col_idx, value)

    for row_idx in range(SCHEDULE_HEADER_ROW, ws.max_row + 1):
        for col_idx in range(3, ws.max_column + 1):
            ws.cell(row_idx, col_idx).alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = f"A{SCHEDULE_DATA_ROW}"
    apply_autofilter(ws, SCHEDULE_HEADER_ROW)
    apply_number_format(ws)
    auto_width(ws)
    ws.column_dimensions[get_column_letter(max_columns)].width = 12

    if schedule_rows:
        total_row = SCHEDULE_DATA_ROW + len(schedule_rows)
        qty_col = max_columns
        ws.cell(total_row, qty_col - 1, "TOTAL")
        ws.cell(total_row, qty_col,
                f"=SUBTOTAL(9,{get_column_letter(qty_col)}{SCHEDULE_DATA_ROW}:{get_column_letter(qty_col)}{total_row - 1})")
        for col_idx in range(1, max_columns + 1):
            c = ws.cell(total_row, col_idx)
            c.font = Font(bold=True)
            c.border = Border(top=Side(style="thin"), bottom=Side(style="double"))

    apply_schedule_row_heights(ws)
    try_insert_logo(ws, logo_bytes)

    return {
        "log_sheet": log_sheet_name,
        "schedule_sheet": sheet_title,
        "matched_codes_count": len(matched_codes),
        "schedule_rows_count": len(schedule_rows),
        "total_matched_qty": total_matched_qty,
        "missing_from_this_log": sorted(c for c in summary.keys() if c not in log_codes),
        "log_warnings": log_warnings,
        "valid_log_columns": valid_columns,
        "scheduled_codes": sorted(matched_codes),
    }

def create_drawing_codes_sheet(wb, reports, summary, descriptions):
    ws = wb.create_sheet("Drawing Codes")
    drawing_map = {}
    for rep in reports:
        log_sheet = rep.get("log_sheet", "")
        for part_code in rep.get("scheduled_codes", []):
            code = clean_code(part_code)
            drawing_code = get_drawing_code(code)
            if code is None or drawing_code is None:
                continue
            if drawing_code not in drawing_map:
                drawing_map[drawing_code] = {"parts": set(), "log_sheets": set(), "total_qty": 0}
            drawing_map[drawing_code]["parts"].add(code)
            drawing_map[drawing_code]["log_sheets"].add(log_sheet)
            drawing_map[drawing_code]["total_qty"] += to_number(summary.get(code, 0))

    rows = [["DRAWING CODE", "PART CODES", "PART COUNT", "TOTAL QTY", "LOG SHEETS"]]
    for drawing_code in sorted(drawing_map.keys()):
        item = drawing_map[drawing_code]
        rows.append([drawing_code, ", ".join(sorted(item["parts"])), len(item["parts"]),
                     item["total_qty"], ", ".join(sorted(item["log_sheets"]))])
    if len(rows) == 1:
        rows.append(["OK", "No scheduled part codes found", 0, 0, ""])

    write_matrix(ws, rows)
    style_header_row(ws, 1, fill_color="D9EAD3")
    ws.freeze_panes = "A2"
    apply_autofilter(ws, 1)
    apply_number_format(ws)
    auto_width(ws, max_width=80)

def create_production_schedule_bytes(log_bytes, selected_log_sheets, summary, descriptions, logo_bytes=None) -> bytes:
    wb = Workbook()
    report_ws = wb.active
    report_ws.title = "Report"
    existing_titles = {"Report"}
    reports = []

    for sheet_name in selected_log_sheets:
        reports.append(create_schedule_sheet_from_log(wb, log_bytes, sheet_name, summary, descriptions, existing_titles, logo_bytes=logo_bytes))

    create_drawing_codes_sheet(wb, reports, summary, descriptions)
    existing_titles.add("Drawing Codes")

    # Report sheet
    report_rows = [["Production Schedule Report"], [],
                   ["Selected LOG sheet", "Schedule sheet", "Matched codes", "Schedule rows", "Total QTY", "Used LOG columns"]]
    for rep in reports:
        report_rows.append([rep["log_sheet"], rep["schedule_sheet"], rep["matched_codes_count"],
                            rep["schedule_rows_count"], rep["total_matched_qty"],
                            ", ".join(str(c) for c in rep["valid_log_columns"])])
    write_matrix(report_ws, report_rows)
    report_ws["A1"].font = Font(bold=True, size=14)
    style_header_row(report_ws, 3)
    apply_number_format(report_ws)
    auto_width(report_ws)

    # Missing in selected LOG
    selected_log_codes = set()
    all_log_warnings = []
    for rep in reports:
        all_log_warnings.extend(rep["log_warnings"])
        _, _, code_to_index, _, _ = read_log_sheet(log_bytes, rep["log_sheet"])
        selected_log_codes.update(code_to_index.keys())

    missing_all = sorted(c for c in summary.keys() if c not in selected_log_codes)
    ws_missing = wb.create_sheet("Missing in Selected LOG")
    rows = [["REF NUMBER", "DESCRIPTION", "TOTAL QTY", "Issue"]]
    for code in missing_all:
        rows.append([code, descriptions.get(code, ""), summary.get(code, 0), "Code in BOM but not in selected LOG sheets"])
    if len(rows) == 1:
        rows.append(["OK", "", "", "All BOM codes found in selected LOG sheets"])
    write_matrix(ws_missing, rows)
    style_header_row(ws_missing, 1, fill_color="F4CCCC")
    ws_missing.freeze_panes = "A2"
    apply_autofilter(ws_missing, 1)
    apply_number_format(ws_missing)
    auto_width(ws_missing)

    # BOM Summary
    ws_summary = wb.create_sheet("BOM Summary")
    rows = [["REF NUMBER", "DESCRIPTION", "TOTAL QTY"]]
    for code in sorted(summary.keys()):
        rows.append([code, descriptions.get(code, ""), summary[code]])
    write_matrix(ws_summary, rows)
    style_header_row(ws_summary, 1)
    ws_summary.freeze_panes = "A2"
    apply_autofilter(ws_summary, 1)
    apply_number_format(ws_summary)
    auto_width(ws_summary)

    # LOG Checks
    ws_warn = wb.create_sheet("LOG Checks")
    rows = [["LOG Sheet", "Row", "Issue", "Value 1", "Value 2", "Value 3"]]
    rows.extend(all_log_warnings)
    if len(rows) == 1:
        rows.append(["OK", "", "No LOG warnings found", "", "", ""])
    write_matrix(ws_warn, rows)
    style_header_row(ws_warn, 1, fill_color="FFF2CC")
    ws_warn.freeze_panes = "A2"
    apply_autofilter(ws_warn, 1)
    auto_width(ws_warn)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ─────────────────────────────────────────────────────────────
# UI
# ─────────────────────────────────────────────────────────────

# Hero
st.markdown("""
<div class="hero-banner">
  <h1>⚙️ BOM Production Automation</h1>
  <p>Upload your BOM and LOG files, select the production units you need,
     and download two ready-to-use Excel reports — in seconds.</p>
  <div class="hero-steps">
    <div class="hero-step"><span class="step-num">1</span> Upload files</div>
    <span class="hero-arrow">→</span>
    <div class="hero-step"><span class="step-num">2</span> Select LOG sheets</div>
    <span class="hero-arrow">→</span>
    <div class="hero-step"><span class="step-num">3</span> Run</div>
    <span class="hero-arrow">→</span>
    <div class="hero-step"><span class="step-num">4</span> Download reports</div>
  </div>
  <div class="trust-badge">🔒 Your files are processed in memory and never stored on any server.</div>
</div>
""", unsafe_allow_html=True)

# ── STEP 1: Upload ───────────────────────────────────────────
st.markdown('<div class="section-header">Step 1 — Upload your files</div>', unsafe_allow_html=True)

st.markdown("""
<div class="info-box">
  <strong>BOM file</strong> — your Bill of Materials Excel spreadsheet (.xlsx / .xlsm).<br>
  Each sheet in the BOM file represents one assembly unit.
</div>
""", unsafe_allow_html=True)
bom_file = st.file_uploader("BOM file (.xlsx / .xlsm)", type=["xlsx", "xlsm"], label_visibility="collapsed")

st.markdown("""
<div class="info-box" style="margin-top:12px">
  <strong>LOG file</strong> — your production tracking Excel spreadsheet (.xlsx / .xlsm).<br>
  This is <em>not</em> a text log — it's an Excel file where each sheet is one production unit (e.g. U-1000, U-2000).
</div>
""", unsafe_allow_html=True)
log_file = st.file_uploader("LOG file (.xlsx / .xlsm)", type=["xlsx", "xlsm"], label_visibility="collapsed")

# ── LOGO ─────────────────────────────────────────────────────
st.markdown('<div class="section-header">Company Logo (optional)</div>', unsafe_allow_html=True)

st.markdown("""
<div class="info-box">
  The logo will appear in the top-left corner (cells A1–A3) of every Production Schedule sheet.
  If you skip this, no logo will be inserted.
</div>
""", unsafe_allow_html=True)

col_logo1, col_logo2 = st.columns([2, 1])

with col_logo1:
    logo_file = st.file_uploader(
        "Upload your logo (.png / .jpg)",
        type=["png", "jpg", "jpeg"],
        label_visibility="collapsed",
        help="Recommended: PNG with transparent background. Will be scaled to ~80px height."
    )

# Load default logo from repo if it exists and user didn't upload one
import os
logo_bytes = None
default_logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Logo.png")

if logo_file is not None:
    logo_bytes = logo_file.read()
    with col_logo2:
        st.image(logo_bytes, caption="Your logo", width=120)
elif os.path.exists(default_logo_path):
    with open(default_logo_path, "rb") as f:
        logo_bytes = f.read()
    with col_logo2:
        st.image(logo_bytes, caption="Default logo (from repo)", width=120)
else:
    with col_logo2:
        st.caption("No logo — reports will have no logo in header.")

# ── STEP 2: Select LOG sheets ────────────────────────────────
if log_file:
    st.markdown('<div class="section-header">Step 2 — Select LOG sheets</div>', unsafe_allow_html=True)

    log_bytes = log_file.read()
    log_file.seek(0)

    try:
        sheet_names = get_log_sheet_names(log_bytes)
        st.markdown("""
        <div class="info-box">
          These are the sheet tabs found in your LOG file. Each sheet is one production unit.
          <strong>Tick the ones you want to include in the Production Schedule.</strong>
        </div>
        """, unsafe_allow_html=True)

        col1, col2 = st.columns([3, 1])
        with col1:
            st.caption(f"Found {len(sheet_names)} sheet(s) in LOG file")
        with col2:
            if st.button("✅ Select all", use_container_width=True):
                st.session_state["select_all"] = True

        selected_sheets = []
        cols = st.columns(3)
        for i, name in enumerate(sheet_names):
            default = st.session_state.get("select_all", True)
            checked = cols[i % 3].checkbox(name, value=default, key=f"sheet_{name}")
            if checked:
                selected_sheets.append(name)

        if selected_sheets:
            st.caption(f"✔ {len(selected_sheets)} sheet(s) selected: {', '.join(selected_sheets)}")
        else:
            st.warning("Please select at least one LOG sheet.")

    except Exception as e:
        st.error(f"Could not read LOG file: {e}")
        selected_sheets = []
        log_bytes = None
else:
    selected_sheets = []
    log_bytes = None

# ── STEP 3: Run ──────────────────────────────────────────────
st.markdown('<div class="section-header">Step 3 — Run</div>', unsafe_allow_html=True)

# Output preview
st.markdown("""
<div style="margin-bottom:16px">
  <div class="output-box">
    <div class="output-box-title">📊 Combined BOM.xlsx</div>
    <div class="output-box-sheets">
      <span class="output-box-tag">Units</span>
      <span class="output-box-tag">Parts</span>
      <span class="output-box-tag">Summary All</span>
      <span class="output-box-tag">Checks</span>
    </div>
  </div>
  <div class="output-box">
    <div class="output-box-title">📅 Production Schedule.xlsx</div>
    <div class="output-box-sheets">
      <span class="output-box-tag">Schedule (per unit)</span>
      <span class="output-box-tag">Drawing Codes</span>
      <span class="output-box-tag">Missing in Selected LOG</span>
      <span class="output-box-tag">BOM Summary</span>
      <span class="output-box-tag">LOG Checks</span>
      <span class="output-box-tag">Report</span>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

run_ready = bom_file is not None and log_bytes is not None and len(selected_sheets) > 0
run_clicked = st.button("⚡  Run", type="primary", use_container_width=True, disabled=not run_ready)

if not run_ready and (bom_file is None or log_file is None):
    st.caption("Upload both files to enable Run.")

# ── STEP 4: Results ──────────────────────────────────────────
if run_clicked and run_ready:
    with st.spinner("Processing your files..."):
        try:
            bom_bytes = bom_file.read()

            # Combined BOM
            combined_bytes, summary, descriptions, bom_warnings = create_combined_workbook_bytes(bom_bytes)

            # Production Schedule
            schedule_bytes = create_production_schedule_bytes(log_bytes, selected_sheets, summary, descriptions, logo_bytes=logo_bytes)

            # ZIP both
            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                zf.writestr("Combined BOM.xlsx", combined_bytes)
                zf.writestr("Production Schedule.xlsx", schedule_bytes)
            zip_buf.seek(0)

            st.success("✅ Done! Your files are ready to download.")
            st.markdown('<div class="section-header">Step 4 — Download your reports</div>', unsafe_allow_html=True)

            col1, col2, col3 = st.columns(3)
            with col1:
                st.download_button(
                    label="📊 Combined BOM.xlsx",
                    data=combined_bytes,
                    file_name="Combined BOM.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            with col2:
                st.download_button(
                    label="📅 Production Schedule.xlsx",
                    data=schedule_bytes,
                    file_name="Production Schedule.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            with col3:
                st.download_button(
                    label="📦 Download both (.zip)",
                    data=zip_buf.getvalue(),
                    file_name="FabTools_Reports.zip",
                    mime="application/zip",
                    use_container_width=True,
                )

            # Summary stats
            with st.expander("📋 Processing summary"):
                st.write(f"**BOM codes found:** {len(summary)}")
                st.write(f"**LOG sheets processed:** {len(selected_sheets)}")
                if bom_warnings:
                    st.warning(f"{len(bom_warnings)} BOM warning(s) — see the Checks sheet in Combined BOM.xlsx")

        except Exception as e:
            st.error(f"❌ Error during processing: {e}")
            st.info("Please check that your files match the expected format. Download sample files from the FabTools homepage.")

# Footer
st.markdown("""
<div class="footer-bar">
  <a href="https://oleksandr-horiachyi.github.io/BOM/" target="_blank">← Back to FabTools</a>
  &nbsp;·&nbsp;
  Built with Python &amp; Streamlit
  &nbsp;·&nbsp;
  <a href="https://github.com/oleksandr-horiachyi/BOM" target="_blank">GitHub</a>
</div>
""", unsafe_allow_html=True)
